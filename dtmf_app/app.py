# -*- coding: utf-8 -*-
"""
DTMF Analyzer - Flask Backend
==============================
API REST que recibe un archivo de audio, lo procesa con el pipeline
DTMF (filtrado, reduccion de ruido, amplificacion, Goertzel) y devuelve
los tonos detectados junto con una imagen del espectrograma.

v2: DTMF engine y configuracion importados desde core/ (sin codigo duplicado).
"""

import os
import sys
import shutil
import uuid
import json
import base64
import subprocess
import traceback
import threading
import csv
import time
import queue as _queue
from collections import deque
from datetime import datetime
from math import gcd

# ── Core DTMF — fuente unica de verdad ────────────────────────
sys.path.insert(0, os.path.dirname(__file__))
from core.config import (
    TARGET_SR, DTMF_MAP, ROW_FREQS, COL_FREQS, DTMF_DIGIT_FREQS,
    FRAME_MS, HOP_MS, MIN_TONE_MS, ENERGY_THRESHOLD, AMPLIFY_DB,
    ROW_DOM_THRESHOLD, COL_DOM_THRESHOLD, TOTAL_DOM_THRESHOLD,
    CONCENTRATION_THRESHOLD, DIGIT_COLORS,
    IVR_DEFAULT_DELAY_S, IVR_DEFAULT_TONE_TIMEOUT, IVR_DEFAULT_MENU_REPEATS,
    IVR_DIAL_TIMEOUT, IVR_MIN_DIALING_SECS, IVR_POST_ACTIVE_LISTEN,
    ADB_WATCHDOG_INTERVAL,
    PRECALL_FRAME_MS, RING_FREQS, RING_E_THR, FLAT_TONE, FLAT_VOICE,
    ZCR_VOICE, RING_ON_MIN, RING_ON_MAX, RING_OFF_MIN,
    VOICE_SUSTAINED_MIN, ENERGY_THR_SIGNAL, ENERGY_SUSTAINED_MIN, MAX_RINGS,
    MONITOR_WINDOW_MS, MONITOR_HOP_MS, MONITOR_VIZ_HZ,
    CALL_MODE_IVR, CALL_MODE_BRIDGE, CALL_MODE_IVR_BRIDGE,
    BRIDGE_TRIGGER_DIGIT_DEFAULT, BRIDGE_BLOCK_MS, BRIDGE_GAIN_IN, BRIDGE_GAIN_OUT,
    CALL_MONITOR_TIMEOUT_S,
)
from core.dtmf_engine import (
    get_bandpass_sos, bandpass_filter, amplify as _amplify_audio,
    detect_dtmf_frame, analyze_dtmf, build_chart, goertzel_batch,
    resample_audio as _resample_audio,
)
from core.audio_bridge import AudioBridge
from core.templates import template_manager, TMPL_AUDIO

# Alias para retrocompatibilidad con el resto del archivo
_DTMF_DIGIT_FREQS = DTMF_DIGIT_FREQS

# IVR — reproducción de audio
try:
    import pygame
    # Inicializar mixer UNA SOLA VEZ al arrancar (no reinicializar entre pistas)
    pygame.mixer.pre_init(frequency=44100, size=-16, channels=1, buffer=1024)
    pygame.mixer.init()
    _PYGAME_OK = True
    print("[OK] pygame.mixer inicializado")
except Exception as _e:
    _PYGAME_OK = False
    print(f"[WARN] pygame no disponible — reproducción de audio desactivada: {_e}")

# IVR — lectura de Excel
try:
    import openpyxl
    _OPENPYXL_OK = True
except Exception:
    _OPENPYXL_OK = False
    print("[WARN] openpyxl no disponible — carga de Excel desactivada")

# IVR — CallMonitor (ahora dentro de dtmf_app/)
sys.path.insert(0, os.path.dirname(__file__))   # asegura imports locales dentro de dtmf_app
try:
    from estado_llamada import CallMonitor
    _MONITOR_OK = True
except Exception as _e:
    _MONITOR_OK = False
    print(f"[WARN] CallMonitor no disponible: {_e}")

# IVR — WhatsApp Notifier (ahora dentro de dtmf_app/notificaciones/)
try:
    from notificaciones.whatsapp_ivr_notifier import (
        WhatsAppIVRNotifier, build_notification_message
    )
    _WA_OK = True
except Exception as _e:
    _WA_OK = False
    print(f"[WARN] WhatsAppIVRNotifier no disponible: {_e}")

# Audio del sistema (monitor DTMF en Python)
try:
    import sounddevice as sd
    _SD_OK = True
except Exception:
    _SD_OK = False
    print("[WARN] sounddevice no disponible")

import numpy as np
import soundfile as sf
import noisereduce as nr
import matplotlib
matplotlib.use("Agg")   # sin GUI - render a buffer
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from scipy.signal import butter, sosfilt, resample_poly
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from flask_socketio import SocketIO, emit

# ──────────────────────────────────────────────────────────────
# Descubrimiento de ffmpeg  (portable — no hardcoded al usuario Milton)
# ──────────────────────────────────────────────────────────────
def _find_ffmpeg() -> str | None:
    """Busca ffmpeg en PATH (shutil.which) y luego en rutas conocidas de Windows."""
    # 1. PATH del sistema (instalación global, winget, conda, etc.)
    found = shutil.which("ffmpeg")
    if found:
        return found
    # 2. Rutas fijas de instalaciones conocidas en Windows
    _candidates = [
        r"C:\Users\Milton\AppData\Local\Microsoft\WinGet\Packages"
        r"\Gyan.FFmpeg_Microsoft.Winget.Source_8wekyb3d8bbwe"
        r"\ffmpeg-8.1-full_build\bin\ffmpeg.exe",
        r"C:\ffmpeg\bin\ffmpeg.exe",
        r"C:\Program Files\ffmpeg\bin\ffmpeg.exe",
        r"C:\ProgramData\chocolatey\bin\ffmpeg.exe",
    ]
    for path in _candidates:
        if os.path.isfile(path):
            return path
    return None

FFMPEG_EXE = _find_ffmpeg()
if FFMPEG_EXE:
    _bin = os.path.dirname(FFMPEG_EXE)
    if _bin not in os.environ.get("PATH", ""):
        os.environ["PATH"] = _bin + os.pathsep + os.environ.get("PATH", "")
    print(f"[OK] ffmpeg: {FFMPEG_EXE}")
else:
    print("[WARN] ffmpeg no encontrado — solo se aceptarán archivos WAV")

# ── Tabla DTMF y parámetros: importados desde core/config.py ──
# (ver core/config.py para calibración y comentarios)

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

ALLOWED_EXTENSIONS = {
    "wav", "mp3", "m4a", "aac", "ogg", "opus",
    "flac", "wma", "mp4", "webm", "3gp", "amr"
}

app = Flask(__name__, static_folder="static", static_url_path="/static")
CORS(app, resources={r"/*": {"origins": "*"}})
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024   # 100 MB max
socketio = SocketIO(app, cors_allowed_origins="*", async_mode="threading")


# ══════════════════════════════════════════════
#  PIPELINE DE AUDIO (helpers app-specific)
#  detect_dtmf_frame, analyze_dtmf, build_chart, bandpass_filter
#  y amplify vienen de core.dtmf_engine (importados arriba).
# ══════════════════════════════════════════════

def convert_to_wav(src_path: str) -> str:
    """Convierte cualquier formato a WAV 8 kHz mono usando ffmpeg."""
    if not FFMPEG_EXE:
        raise RuntimeError("ffmpeg no disponible para convertir este formato.")
    dst_path = src_path + "_conv.wav"
    result = subprocess.run(
        [FFMPEG_EXE, "-y", "-i", src_path,
         "-ar", str(TARGET_SR), "-ac", "1", "-f", "wav", dst_path],
        capture_output=True, text=True, timeout=120
    )
    if not os.path.isfile(dst_path):
        raise RuntimeError(f"ffmpeg falló: {result.stderr[-500:]}")
    return dst_path


def load_audio(path: str):
    """
    Carga el audio como float32 normalizado en [-1, 1].
    Convierte a WAV si es necesario y resamplea a TARGET_SR.
    Retorna (audio, sr, path_cargado).
    """
    ext = os.path.splitext(path)[1].lower().lstrip(".")
    if ext != "wav":
        path = convert_to_wav(path)

    audio, sr = sf.read(path, always_2d=False)
    if audio.ndim > 1:
        audio = audio.mean(axis=1)
    audio = audio.astype(np.float32)

    if sr != TARGET_SR:
        audio = _resample_audio(audio, sr, TARGET_SR)
        sr = TARGET_SR

    return audio, sr, path   # path puede haber cambiado si se convirtió


def reduce_noise_audio(audio: np.ndarray, sr: int) -> np.ndarray:
    """Reducción de ruido con noisereduce (perfil del primer 10% de la señal)."""
    noise_len = max(int(0.1 * len(audio)), sr // 2)
    noise_clip = audio[:noise_len]
    return nr.reduce_noise(
        y=audio, y_noise=noise_clip, sr=sr,
        stationary=False, prop_decrease=0.85
    ).astype(np.float32)


def amplify(audio: np.ndarray, gain_db: float = AMPLIFY_DB) -> np.ndarray:
    """Wrapper local de amplify para retrocompatibilidad con llamadas directas en este módulo."""
    return _amplify_audio(audio, gain_db)


def _get_bandpass(sr: int) -> np.ndarray:
    """Acceso al caché de filtro pasa-banda desde core.dtmf_engine."""
    return get_bandpass_sos(sr)


# ══════════════════════════════════════════════
#  RUTAS FLASK
# ══════════════════════════════════════════════

@app.route("/")
def index():
    return send_from_directory("static", "index.html")


@app.route("/analyze", methods=["POST"])
def analyze():
    if "audio" not in request.files:
        return jsonify({"error": "No se recibio ningun archivo."}), 400

    f = request.files["audio"]
    if not f or not f.filename:
        return jsonify({"error": "Archivo invalido."}), 400

    ext = os.path.splitext(f.filename)[1].lower().lstrip(".")
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({"error": f"Formato '{ext}' no soportado."}), 400

    # Guardar con nombre seguro
    safe_name = f"{uuid.uuid4().hex}.{ext}"
    src_path  = os.path.join(UPLOAD_FOLDER, safe_name)
    f.save(src_path)

    tmp_files = [src_path]
    try:
        # Pipeline
        audio, sr, loaded_path = load_audio(src_path)
        if loaded_path != src_path:
            tmp_files.append(loaded_path)

        duration = len(audio) / sr

        audio = bandpass_filter(audio, sr)
        audio = reduce_noise_audio(audio, sr)
        audio = amplify(audio, AMPLIFY_DB)

        tones = analyze_dtmf(audio, sr)
        sequence = "".join(t["digit"] for t in tones)

        chart_b64 = build_chart(audio, sr, tones, duration)

        payload = {
            "ok"        : True,
            "filename"  : f.filename,
            "duration_s": round(duration, 2),
            "sample_rate": sr,
            "tones"     : tones,
            "sequence"  : sequence,
            "chart"     : chart_b64,
        }
        return jsonify(payload)

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

    finally:
        for p in tmp_files:
            try:
                os.remove(p)
            except Exception:
                pass


@app.route("/health")
def health():
    return jsonify({"status": "ok", "ffmpeg": FFMPEG_EXE or "not found"})


# ══════════════════════════════════════════════
#  WEBSOCKET — MONITOR EN TIEMPO REAL
# ══════════════════════════════════════════════

# Buffer deslizante por sesion (sid -> list of float32)
_rt_buffers = {}
_rt_sr       = {}   # sample rate del cliente por sesion

# Filtro pasa-banda reutilizable (se construye una vez por sr)
_bp_cache = {}

def _get_bandpass(sr: int):
    if sr not in _bp_cache:
        nyq = sr / 2.0
        sos = butter(4, [300 / nyq, 3400 / nyq], btype="band", output="sos")
        _bp_cache[sr] = sos
    return _bp_cache[sr]


@socketio.on("connect")
def on_connect():
    sid = request.sid
    _rt_buffers[sid] = []
    _rt_sr[sid]      = 8000
    emit("connected", {"sid": sid})


@socketio.on("disconnect")
def on_disconnect():
    sid = request.sid
    _rt_buffers.pop(sid, None)
    _rt_sr.pop(sid, None)


@socketio.on("rt_config")
def on_rt_config(data):
    """Cliente informa su sample rate."""
    sid = request.sid
    _rt_sr[sid] = int(data.get("sampleRate", 44100))


@socketio.on("audio_chunk")
def on_audio_chunk(data):
    """
    Recibe un chunk de PCM Float32 del navegador (base64 o lista),
    lo acumula en un buffer deslizante de 80 ms,
    aplica Goertzel y emite el digito detectado (o None).
    """
    sid       = request.sid
    client_sr = _rt_sr.get(sid, 44100)

    # Decodificar: el cliente envia JSON con {pcm: [f32, ...], sr: int}
    pcm_list  = data.get("pcm", [])
    if not pcm_list:
        return

    client_sr = int(data.get("sr", client_sr))
    _rt_sr[sid] = client_sr

    # Convertir a numpy float32
    chunk = np.array(pcm_list, dtype=np.float32)

    # Resamplear a 8 kHz si es necesario
    if client_sr != TARGET_SR:
        g     = gcd(TARGET_SR, client_sr)
        chunk = resample_poly(chunk, TARGET_SR // g, client_sr // g).astype(np.float32)

    # Acumular en buffer
    buf = _rt_buffers.setdefault(sid, [])
    buf.extend(chunk.tolist())

    # Ventana de analisis: 80 ms = 640 muestras a 8kHz
    WINDOW = int(TARGET_SR * 0.08)   # 640
    HOP    = int(TARGET_SR * 0.02)   # 160 (20 ms hop)

    if len(buf) < WINDOW:
        return   # todavia no hay suficientes muestras

    # Tomar la ventana mas reciente
    frame_np = np.array(buf[-WINDOW:], dtype=np.float32)

    # Filtro pasa-banda
    sos        = _get_bandpass(TARGET_SR)
    frame_filt = sosfilt(sos, frame_np).astype(np.float32)

    energy = float(np.mean(frame_filt ** 2))
    if energy < ENERGY_THRESHOLD:
        emit("rt_digit", {"digit": None, "energy": 0.0})
        # Limpiar buffer acumulado (silencio)
        _rt_buffers[sid] = []
        return

    digit = detect_dtmf_frame(frame_filt, TARGET_SR, energy)

    # Mantener buffer deslizante: descartar muestras antiguas
    if len(buf) > WINDOW * 4:
        _rt_buffers[sid] = buf[-WINDOW:]

    emit("rt_digit", {
        "digit" : digit,
        "energy": round(float(energy), 8),
    })

    # Diagnostico: si la campaña esta activa, contar frames recibidos
    if _ivr_dtmf_callback:
        if not hasattr(on_audio_chunk, "_cnt"): on_audio_chunk._cnt = 0
        on_audio_chunk._cnt += 1
        if on_audio_chunk._cnt % 25 == 0: # aprox cada 1 seg (40ms * 25)
            _emit_ivr("ivr_log", {"msg": f"  [DEBUG] Mic activo ({energy:.2e})", "level": "info"})

    # Si hay una campaña IVR activa y se detectó un dígito real → notificarla
    if digit and _ivr_dtmf_callback:
        print(f"[IVR-RT] Digito '{digit}' detectado (E={energy:.2e}) -> enviando a campaña")
        try:
            _ivr_dtmf_callback(digit)
        except Exception as exc:
            print(f"[IVR-RT] Error en callback: {exc}")


# ══════════════════════════════════════════════
#  IVR AUTOMATOR — Configuración y estado global
# ══════════════════════════════════════════════

IVR_RESULTS_CSV    = os.path.join(os.path.dirname(__file__), "ivr_results.csv")
IVR_AUDIO_FOLDER   = os.path.join(os.path.dirname(__file__), "ivr_audio")
IVR_RECORDINGS_DIR = os.path.join(os.path.dirname(__file__), "recordings")
os.makedirs(IVR_AUDIO_FOLDER,   exist_ok=True)
os.makedirs(IVR_RECORDINGS_DIR, exist_ok=True)

# Estado global de la campaña (singleton)
_ivr_campaign: "IVRCampaign | None" = None
_ivr_lock = threading.Lock()
_adb_watchdog: "ADBWatchdog | None" = None   # monitor de conexión ADB

# Cuando el IVR está ACTIVO, cualquier dígito DTMF detectado por
# el monitor de micrófono se desvía aquí en lugar de sólo emitirse a la UI
_ivr_dtmf_callback = None   # callable(digit) | None

# Dispositivo de salida de audio seleccionado (nombre para pygame)
_audio_output_device_name:  str | None = None
_audio_output_device_index: int | None = None   # índice del dispositivo de salida (loopback)
_active_recorder: "CallRecorder | None" = None  # grabador activo — recibe audio del monitor

# ── WhatsApp Notifications ──────────────────────────────────────────
WA_CONFIG_FILE = os.path.join(os.path.dirname(__file__), "wa_notif_config.json")

# Instancia única del notificador
_wa_notifier: "WhatsAppIVRNotifier | None" = None

# Config activa (se carga del JSON al iniciar)
_wa_config: dict = {
    "enabled": False,
    "contact": "",   # grupo o número principal
    "backup":  "",   # número de respaldo
}


def _wa_load_config():
    """Carga la configuración de notificaciones WA desde disco."""
    global _wa_config, _wa_notifier
    if os.path.isfile(WA_CONFIG_FILE):
        try:
            with open(WA_CONFIG_FILE, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            _wa_config.update(data)
            print(f"[WA] Config cargada: enabled={_wa_config['enabled']} contact='{_wa_config['contact']}'")
        except Exception as exc:
            print(f"[WA] Error leyendo config: {exc}")
    # Inicializar instancia del notificador
    if _WA_OK:
        _wa_notifier = WhatsAppIVRNotifier()


def _emit_output_viz(path: str, cancel_event: threading.Event = None):
    """
    Hilo auxiliar: lee el audio del archivo y emite audio_viz ch='output'
    a ~15 Hz sincronizado con la duración real del archivo.
    Se ejecuta en paralelo a pygame.mixer para no bloquear la reproducción.
    """
    try:
        import wave as _wave
        import struct
        try:
            import soundfile as _sf
            data, sr = _sf.read(path, dtype="float32", always_2d=False)
            if data.ndim > 1:
                data = data[:, 0]
        except Exception:
            # Fallback: onda sin importar soundfile (solo WAV)
            try:
                with _wave.open(path, "rb") as wf:
                    sr     = wf.getframerate()
                    n_ch   = wf.getnchannels()
                    sw     = wf.getsampwidth()
                    raw    = wf.readframes(wf.getnframes())
                fmt    = {1: "b", 2: "h", 4: "i"}.get(sw, "h")
                samps  = struct.unpack(f"{len(raw)//sw}{fmt}", raw)
                data   = np.array(samps[::n_ch], dtype=np.float32)
                data  /= float(2 ** (8 * sw - 1))  # normalizar a -1..1
            except Exception:
                return   # no podemos leer el archivo, salir sin viz

        block_size = max(1, sr // 15)   # bloque de ~67 ms → ~15 Hz
        for i in range(0, len(data), block_size):
            if cancel_event and cancel_event.is_set():
                break
            chunk = data[i : i + block_size]
            rms   = float(np.sqrt(np.mean(chunk ** 2))) if len(chunk) else 0.0
            socketio.emit("audio_viz", {"ch": "output", "rms": rms})
            # Alimentar grabador activo con el audio del archivo IVR
            if _active_recorder is not None and len(chunk) > 0:
                _active_recorder.feed_output(chunk.astype(np.float32), sr)
            time.sleep(block_size / sr)    # esperar el tiempo real del bloque
    except Exception as exc:
        print(f"[OutputViz] {exc}")


def _wa_save_config():
    """Persiste la configuración de notificaciones WA en disco."""
    try:
        with open(WA_CONFIG_FILE, "w", encoding="utf-8") as fh:
            json.dump(_wa_config, fh, ensure_ascii=False, indent=2)
    except Exception as exc:
        print(f"[WA] Error guardando config: {exc}")


def _send_whatsapp_notification(number: str, status: str,
                                 digit: str | None = None,
                                 option_desc: str | None = None):
    """
    Envía notificación WhatsApp al finalizar una llamada.
    Se llama desde _process_number() si las notificaciones están activas.
    """
    if not _WA_OK or not _wa_notifier:
        return
    if not _wa_config.get("enabled"):
        return
    contacto = _wa_config.get("contact", "").strip()
    if not contacto:
        print("[WA] ⚠ Sin contacto destino configurado")
        return
    backup = _wa_config.get("backup", "").strip() or None
    mensaje = build_notification_message(number, status, digit, option_desc)
    _wa_notifier.enqueue_notification(contacto, mensaje, backup)


# Cargar config al arrancar
_wa_load_config()


def _save_call_result(number: str, status: str, digit: str | None, notes: str = ""):
    """Guarda el resultado de una llamada en el CSV de resultados."""
    file_exists = os.path.isfile(IVR_RESULTS_CSV)
    with open(IVR_RESULTS_CSV, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(["timestamp", "numero", "estado", "tono_detectado", "notas"])
        writer.writerow([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            number, status, digit or "", notes
        ])


# Nombre del dispositivo de salida activo al iniciar la campaña (para reinit si cambia)
_mixer_device_name: str | None = None


def _ensure_mixer(device_name: str | None = None) -> bool:
    """
    Garantiza que pygame.mixer está inicializado con el dispositivo correcto.
    Solo reinicializa si el dispositivo cambia (evita clics entre pistas consecutivas).
    Retorna True si el mixer está listo.
    """
    global _mixer_device_name
    if not _PYGAME_OK:
        return False
    try:
        current_init = pygame.mixer.get_init()
        # Reinicializar solo si cambia el dispositivo de salida
        if current_init and _mixer_device_name == device_name:
            return True   # ya está en el dispositivo correcto, nada que hacer

        if current_init:
            pygame.mixer.quit()

        if device_name:
            try:
                pygame.mixer.init(devicename=device_name)
                _mixer_device_name = device_name
                print(f"[Mixer] Inicializado en dispositivo: {device_name}")
            except Exception as ex:
                print(f"[Mixer] Fallo dispositivo '{device_name}': {ex} — usando default")
                pygame.mixer.init()
                _mixer_device_name = None
        else:
            pygame.mixer.init()
            _mixer_device_name = None
        return True
    except Exception as exc:
        print(f"[Mixer] Error de inicialización: {exc}")
        return False


def _play_audio(path: str, cancel_event: threading.Event = None):
    """
    Reproduce un archivo de audio usando pygame.
    El mixer solo se reinicializa si el dispositivo de salida cambió desde la
    última llamada (evita clics y crashes entre pistas consecutivas).
    Si cancel_event se activa, detiene la reproducción inmediatamente.
    """
    if not _PYGAME_OK or not path or not os.path.isfile(path):
        return
    try:
        print(f"[IVR-Audio] Reproduciendo: {os.path.basename(path)}")

        if not _ensure_mixer(_audio_output_device_name):
            print("[IVR-Audio] Mixer no disponible — omitiendo audio")
            return

        pygame.mixer.music.load(path)
        pygame.mixer.music.play()

        # Hilo paralelo: waveform de salida sincronizada con el archivo
        threading.Thread(
            target=_emit_output_viz,
            args=(path, cancel_event),
            daemon=True, name="OutputViz"
        ).start()

        start = time.time()
        while pygame.mixer.music.get_busy() and (time.time() - start) < 60:
            if cancel_event and cancel_event.is_set():
                pygame.mixer.music.stop()
                print(f"[IVR-Audio] Reproducción cancelada: {os.path.basename(path)}")
                return
            time.sleep(0.05)   # 50ms poll — antes era 100ms (más responsivo al cuelgue)
        print(f"[IVR-Audio] Fin: {os.path.basename(path)}")
    except Exception as exc:
        print(f"[IVR] Error reproduciendo audio: {exc}")



def _emit_ivr(event: str, data: dict):
    """Emite un evento Socket.IO desde cualquier hilo."""
    socketio.emit(event, data)


# ════════════════════════════════════════════
#  ADB WATCHDOG — Monitor de conexión en tiempo real
# ════════════════════════════════════════════

class ADBWatchdog(threading.Thread):
    """
    Monitorea en tiempo real si el dispositivo ADB sigue conectado.
    - Si se desconecta: pausa la campaña activa y emite alerta a la UI.
    - Si reconecta:     reanuda automáticamente la campaña.
    """
    CHECK_INTERVAL = 3.0   # segundos entre verificaciones

    def __init__(self, device_id: str):
        super().__init__(daemon=True, name=f"ADBWatchdog-{device_id}")
        self.device_id  = device_id
        self._stop_ev   = threading.Event()
        self._connected = True   # optimista al inicio

    # ── Control ──────────────────────────────────────────────────

    def stop(self):
        self._stop_ev.set()

    # ── Hilo principal ────────────────────────────────────────────

    def run(self):
        print(f"[ADBWatchdog] Iniciando monitoreo: {self.device_id}")
        # Primera verificación inmediata
        self._connected = self._check()
        _emit_ivr("adb_status", {
            "connected": self._connected,
            "device_id": self.device_id,
        })

        while not self._stop_ev.wait(self.CHECK_INTERVAL):
            connected = self._check()
            if connected != self._connected:
                self._connected = connected
                self._on_change(connected)

        print(f"[ADBWatchdog] Detenido: {self.device_id}")

    def _check(self) -> bool:
        """Retorna True si el dispositivo responde como 'device' en ADB."""
        try:
            cmd = ["adb", "-s", self.device_id, "get-state"]
            r   = subprocess.run(cmd, capture_output=True, text=True, timeout=5)
            return r.returncode == 0 and "device" in r.stdout
        except Exception:
            return False

    def _on_change(self, connected: bool):
        """Reacciona al cambio de estado de conexión."""
        _emit_ivr("adb_status", {"connected": connected, "device_id": self.device_id})

        if connected:
            # Reconectado — reanudar campaña si estaba pausada
            _emit_ivr("ivr_log", {
                "msg": f"✅ Dispositivo ADB reconectado: {self.device_id}",
                "level": "success"
            })
            if _ivr_campaign and _ivr_campaign.is_running:
                _ivr_campaign.resume()
                _emit_ivr("ivr_log", {"msg": "▶️ Campaña reanudada", "level": "success"})
        else:
            # Desconectado — pausar campaña
            _emit_ivr("ivr_log", {
                "msg": f"⚠️ ADB desconectado: {self.device_id} — campaña en pausa",
                "level": "warn"
            })
            if _ivr_campaign and _ivr_campaign.is_running:
                _ivr_campaign.pause()
                _emit_ivr("ivr_log", {
                    "msg": "⏸️ Esperando reconexión del dispositivo…",
                    "level": "warn"
                })


# ══════════════════════════════════════════════
#  CALL RECORDER — WAV 8 kHz / mono / 16-bit
#  Sin dependencias extra (~1 MB/min de voz)
# ══════════════════════════════════════════════

import wave as _wave

class CallRecorder(threading.Thread):
    """
    Graba el audio de una llamada en WAV estéreo a tasa nativa del dispositivo.

    Canal IZQUIERDO (L) — entrada : micrófono / interfaz USB (voz entrante)
    Canal DERECHO   (R) — salida  : audio IVR reproducido (de _emit_output_viz)

    Cada chunk de salida lleva su propio timestamp absolute — al mezclar se
    reconstruye la línea de tiempo exacta del canal R, preservando:
      • El offset inicial (tiempo desde marcado hasta primer audio IVR)
      • Los silencios entre audios (espera de DTMF, pausas del IVR, etc.)
    Tamaño aproximado: estéreo 44.1 kHz ≈ 10 MB/min.
    """

    def __init__(self, filepath: str):
        super().__init__(daemon=True, name="CallRecorder")
        self.filepath   = filepath
        self._stop_ev   = threading.Event()
        self._in_queue  = _queue.Queue()   # (chunk, sr)  — desde mic
        self._out_queue = _queue.Queue()   # (chunk, t_abs, sr) — desde IVR
        self._sr_in     = None
        self._t0_in     = None             # timestamp primer chunk de mic

    def stop(self):
        self._stop_ev.set()

    def feed(self, chunk: np.ndarray, sr: int):
        """Audio de ENTRADA (micrófono) — desde PythonAudioMonitor o PreCallAnalyzer."""
        if self._sr_in is None:
            self._sr_in = sr
        if self._t0_in is None:
            self._t0_in = time.time()
        try:
            self._in_queue.put_nowait(chunk.copy())
        except _queue.Full:
            pass

    def feed_output(self, chunk: np.ndarray, sr: int):
        """
        Audio de SALIDA (IVR reproduciendo) — desde _emit_output_viz.
        Incluye timestamp absoluto para reconstruir la línea de tiempo exacta,
        preservando los silencios entre archivos IVR.
        """
        try:
            self._out_queue.put_nowait((chunk.copy(), time.time(), sr))
        except _queue.Full:
            pass

    # ──────────────────────────────────────────────────────────────────
    def run(self):
        # Esperar tasa nativa del mic
        deadline = time.time() + 5.0
        while self._sr_in is None and not self._stop_ev.is_set():
            if time.time() > deadline:
                print("[Recorder] Timeout esperando SR del monitor")
                return
            time.sleep(0.05)

        SR         = self._sr_in
        frames_in: list[np.ndarray]                 = []
        frames_out: list[tuple[np.ndarray, float]]  = []  # (chunk_resampled, t_abs)

        def _consume_out_queue(q):
            """Vacía la cola de salida resampleando al SR de entrada."""
            while True:
                try:
                    chunk_o, t_abs, sr_o = q.get_nowait()
                    if sr_o != SR:
                        g       = gcd(SR, sr_o)
                        chunk_o = resample_poly(
                            chunk_o.astype(np.float64), SR // g, sr_o // g
                        ).astype(np.float32)
                    frames_out.append((chunk_o, t_abs))
                except _queue.Empty:
                    break

        # Consumir ambas colas mientras grabamos
        while not self._stop_ev.is_set():
            try:
                chunk = self._in_queue.get(timeout=0.05)
                frames_in.append(chunk)
            except _queue.Empty:
                pass
            _consume_out_queue(self._out_queue)

        # Vaciar restos al detenerse
        while True:
            try:
                frames_in.append(self._in_queue.get_nowait())
            except _queue.Empty:
                break
        _consume_out_queue(self._out_queue)

        if not frames_in:
            print("[Recorder] Sin datos de audio — grabación cancelada")
            return

        # ── Construir canal L (entrada) ────────────────────────────
        audio_in = np.concatenate(frames_in).astype(np.float32)
        n_in     = len(audio_in)

        # ── Construir canal R (salida) con timeline exacta ─────────
        if frames_out and self._t0_in is not None:
            # Determinar la longitud total necesaria para el canal R
            # = posición más tardía (en muestras) + duración del último chunk
            max_end = 0
            positioned: list[tuple[int, np.ndarray]] = []
            for chunk_o, t_abs in frames_out:
                t_offset   = max(0.0, t_abs - self._t0_in)   # segundos desde inicio
                sample_pos = int(t_offset * SR)
                end_pos    = sample_pos + len(chunk_o)
                positioned.append((sample_pos, chunk_o))
                max_end = max(max_end, end_pos)

            # Canal R: ceros (silencio) de longitud suficiente
            n_out    = max(n_in, max_end)
            ch_out   = np.zeros(n_out, dtype=np.float32)

            # Colocar cada chunk en su posición temporal exacta
            for sample_pos, chunk_o in positioned:
                end = sample_pos + len(chunk_o)
                if end <= len(ch_out):
                    ch_out[sample_pos:end] += chunk_o
                else:
                    # Por si acaso el canal R se excede
                    ch_out = np.pad(ch_out, (0, end - len(ch_out)))
                    ch_out[sample_pos:end] += chunk_o

            # Igualar longitudes entre canales L y R
            n      = max(n_in, len(ch_out))
            ch_in  = np.pad(audio_in, (0, n - n_in),        constant_values=0.0)
            ch_out = np.pad(ch_out,   (0, n - len(ch_out)), constant_values=0.0)

            dur_out = len(frames_out)
            print(f"[Recorder] Timeline salida: {dur_out} chunks posicionados sobre {n/SR:.1f}s")
        else:
            # Sin audio IVR → canal R en silencio
            ch_in  = audio_in
            ch_out = np.zeros_like(audio_in)

        # ── Escribir WAV estéreo ───────────────────────────────────
        stereo = np.stack([ch_in, ch_out], axis=1)
        pcm    = np.clip(stereo, -1.0, 1.0)
        pcm    = (pcm * 32767).astype(np.int16)

        os.makedirs(os.path.dirname(self.filepath), exist_ok=True)
        with _wave.open(self.filepath, "wb") as wf:
            wf.setnchannels(2)
            wf.setsampwidth(2)
            wf.setframerate(SR)
            wf.writeframes(pcm.tobytes())

        size_kb = os.path.getsize(self.filepath) / 1024
        mode    = "estéreo (mic+IVR timeline)" if frames_out else "estéreo (mic+silencio)"
        print(f"[Recorder] Guardado: {os.path.basename(self.filepath)} ({size_kb:.0f} KB, {mode} @ {SR} Hz)")








# ══════════════════════════════════════════════
#  PRE-CALL AUDIO ANALYZER
#  Detecta ring tones y voz del operador ANTES de que la llamada entre en ACTIVE
#  Técnicas: Goertzel @ ring freqs + Spectral Flatness + ZCR
# ══════════════════════════════════════════════

class PreCallAudioAnalyzer(threading.Thread):
    """
    Analiza el audio capturado internamente durante la fase DIALING/CONNECTING
    para distinguir:
      - Tonos de timbre (ring tones): señal tonal estrecha con patrón ON/OFF
      - Voz del operador: señal broadband sostenida (número apagado/no disponible)

    Resultado:
      self.ring_count     (int)  — número de rings completos detectados
      self.operator_voice (bool) — True si se detectó voz del operador

    Todos los parámetros se toman de core.config (fuente única de verdad).
    """

    # ── Parámetros — desde core.config (no duplicar aquí) ────────────────────
    FRAME_MS          = PRECALL_FRAME_MS       # ms por ventana de análisis
    RING_FREQS        = RING_FREQS             # Hz — Colombia: ~425 Hz
    RING_E_THR        = RING_E_THR             # energía mínima para procesar el frame
    FLAT_TONE         = FLAT_TONE              # spectral flatness ≤ → frame tonal
    FLAT_VOICE        = FLAT_VOICE             # spectral flatness ≥ → frame broadband
    ZCR_VOICE         = ZCR_VOICE              # ZCR normalizado ≥ → componente vocal
    RING_ON_MIN       = RING_ON_MIN            # s — duración mínima burst de ring
    RING_ON_MAX       = RING_ON_MAX            # s — duración máxima burst de ring
    RING_OFF_MIN      = RING_OFF_MIN           # s — silencio mínimo entre rings
    VOICE_SUSTAINED_MIN = VOICE_SUSTAINED_MIN  # s de voz continua → operador
    ENERGY_THR_SIGNAL = ENERGY_THR_SIGNAL      # RMS mínimo para 'hay señal'
    ENERGY_SUSTAINED_MIN = ENERGY_SUSTAINED_MIN # s de señal sostenida → operador
    MAX_RINGS         = MAX_RINGS              # límite de rings para UNAVAILABLE

    def __init__(self, device_index=None):
        super().__init__(daemon=True, name="PreCallAudioAnalyzer")
        self.device_index   = device_index
        self._stop_ev       = threading.Event()
        self.ring_count     = 0
        self.operator_voice = False
        # Estado interno del detector de ring
        self._in_ring       = False
        self._ring_start    = 0.0
        self._silence_start = 0.0
        self._in_silence    = False
        self._pending_ring  = False
        self._pending_start = 0.0
        # Contador de frames de voz CONSECUTIVOS (detector espectral)
        self._consecutive_voice_frames  = 0
        # Contador de frames con ENERGÍA sostenida (detector por energía)
        self._consecutive_energy_frames = 0

    # ── API pública ────────────────────────────────────────────────

    def stop(self):
        self._stop_ev.set()

    def reset_vad_counter(self):
        """Reinicia contadores de voz y energía (llamar al inicio de la ventana post-ACTIVE)."""
        self._consecutive_voice_frames  = 0
        self._consecutive_energy_frames = 0

    # ── Métodos de análisis espectral ──────────────────────────────

    @staticmethod
    def _goertzel_ring_energy(frame: np.ndarray, sr: int) -> float:
        """Suma de energía Goertzel en las frecuencias típicas de ring tone."""
        N = len(frame)
        total = 0.0
        for freq in PreCallAudioAnalyzer.RING_FREQS:
            k      = int(0.5 + N * freq / sr)
            omega  = 2.0 * np.pi * k / N
            coeff  = 2.0 * np.cos(omega)
            s1 = s2 = 0.0
            for x in frame:
                s  = float(x) + coeff * s1 - s2
                s2 = s1
                s1 = s
            total += (s2**2 + s1**2 - coeff * s1 * s2) / (N * N)
        return total

    @staticmethod
    def _spectral_flatness(frame: np.ndarray) -> float:
        """Spectral flatness (Wiener entropy): 0=tonal puro, 1=ruido blanco."""
        fft_mag = np.abs(np.fft.rfft(frame))
        fft_mag = fft_mag[fft_mag > 1e-12]   # evitar log(0)
        if len(fft_mag) < 4:
            return 0.0
        geo_mean = np.exp(np.mean(np.log(fft_mag)))
        ari_mean = np.mean(fft_mag)
        return float(geo_mean / (ari_mean + 1e-12))

    @staticmethod
    def _zcr(frame: np.ndarray) -> float:
        """Zero-crossing rate normalizado (0–1)."""
        signs = np.sign(frame)
        signs[signs == 0] = 1
        crossings = np.sum(np.abs(np.diff(signs))) / 2
        return float(crossings / len(frame))

    # ── Clasificador de frame ──────────────────────────────────────

    def _classify_frame(self, energy: float, ring_e: float,
                        flatness: float, zcr: float) -> str:
        """Clasifica un frame de audio como 'ring', 'voice' o 'silence'."""
        if energy < self.RING_E_THR:
            return "silence"

        # Tonal + energía concentrada en frecuencias de ring → ring
        if flatness <= self.FLAT_TONE and ring_e > energy * 0.30:
            return "ring"

        # Broadband + ZCR alto → voz
        if flatness >= self.FLAT_VOICE and zcr >= self.ZCR_VOICE:
            return "voice"

        # Broadband sin ZCR suficiente pero claramente no tonal → voice (conservador)
        if flatness >= self.FLAT_VOICE:
            return "voice"

        return "silence"   # ambiguo → tratar como silencio

    # ── Máquina de estados para ring detection ─────────────────────

    def _update_ring_state(self, cls: str, now: float):
        """Actualiza el conteo de rings en base a la secuencia ON/OFF."""
        is_active = (cls == "ring")

        if is_active:
            if not self._in_ring:
                # Inicio de nuevo burst
                self._in_ring    = True
                self._ring_start = now
                self._in_silence = False
                self._silence_start = 0.0
        else:
            if self._in_ring:
                burst_dur = now - self._ring_start
                self._in_ring = False
                if self.RING_ON_MIN <= burst_dur <= self.RING_ON_MAX:
                    # Burst de duración correcta → esperamos el silencio para confirmar
                    self._pending_ring  = True
                    self._pending_start = now
                elif burst_dur > self.RING_ON_MAX:
                    # Demasiado largo para ser ring → ignorar
                    self._pending_ring = False
                self._silence_start = now
                self._in_silence    = True

            if self._in_silence and self._pending_ring:
                silence_dur = now - self._pending_start
                if silence_dur >= self.RING_OFF_MIN:
                    # Ring confirmado (burst + silencio correcto)
                    self.ring_count   += 1
                    self._pending_ring = False
                    print(f"[PreCall] Ring #{self.ring_count} detectado")

    # ── Evaluador de voz sostenida (VAD) ───────────────────────────

    def _update_vad(self, cls: str):
        """
        [Método 1 — espectral] Detecta voz SOSTENIDA contando frames consecutivos de 'voice'.
        Un humano dice "Hola?" (~0.5-1.5 s) y pausa; el operador habla sin pausa 2+ s.
        """
        if cls == "voice":
            self._consecutive_voice_frames += 1
        else:
            self._consecutive_voice_frames = 0

        sustained_secs = self._consecutive_voice_frames * self.FRAME_MS / 1000.0
        if self._consecutive_voice_frames > 0 and self._consecutive_voice_frames % 10 == 0:
            print(f"[PreCall] Espectral voz consecutiva: {sustained_secs:.1f}s")
        if sustained_secs >= self.VOICE_SUSTAINED_MIN:
            self.operator_voice = True

    def _update_energy_vad(self, energy: float):
        """
        [Método 2 — energía] Detecta señal CONTINUA por RMS sostenido.
        Funciona independientemente del tipo de señal (voz, música, tono).
        El operador/buzon genera energía constante; el silencio de apagado = RMS ≈ 0.
        RING_OFF cancela el contador (los timbres tienen gaps de silencio).
        """
        rms = float(np.sqrt(energy)) if energy > 0 else 0.0
        if rms >= self.ENERGY_THR_SIGNAL:
            self._consecutive_energy_frames += 1
        else:
            self._consecutive_energy_frames = 0

        sustained_secs = self._consecutive_energy_frames * self.FRAME_MS / 1000.0
        if self._consecutive_energy_frames > 0 and self._consecutive_energy_frames % 10 == 0:
            print(f"[PreCall] Energía sostenida: {sustained_secs:.1f}s  RMS={rms:.4f}")
        if sustained_secs >= self.ENERGY_SUSTAINED_MIN:
            self.operator_voice = True

    # ── Hilo principal ────────────────────────────────────────────

    def run(self):
        if not _SD_OK:
            print("[PreCall] sounddevice no disponible — análisis pre-llamada omitido")
            return
        try:
            dev_info  = sd.query_devices(self.device_index, "input")
            sr_native = int(dev_info["default_samplerate"])
        except Exception as exc:
            print(f"[PreCall] Dispositivo inválido: {exc}")
            return

        frame_samples_native = int(sr_native * self.FRAME_MS / 1000)
        frame_samples_8k     = int(TARGET_SR * self.FRAME_MS / 1000)

        print(f"[PreCall] Iniciado — dispositivo idx={self.device_index} @ {sr_native} Hz")

        # ── Abrir stream con retry (WASAPI puede tardar en liberar el device) ──
        max_retries = 4
        retry_delay = 0.3
        last_exc    = None

        for attempt in range(1, max_retries + 1):
            if self._stop_ev.is_set():
                break
            try:
                with sd.InputStream(
                    device     = self.device_index,
                    channels   = 1,
                    samplerate = sr_native,
                    blocksize  = frame_samples_native,
                    dtype      = "float32",
                ) as stream:
                    while not self._stop_ev.is_set():
                        data, _ = stream.read(frame_samples_native)
                        if self._stop_ev.is_set():
                            break

                        chunk = data[:, 0].astype(np.float32)

                        # Resamplear a 8 kHz para análisis
                        if sr_native != TARGET_SR:
                            g     = gcd(TARGET_SR, sr_native)
                            chunk = resample_poly(
                                chunk, TARGET_SR // g, sr_native // g
                            ).astype(np.float32)

                        # Tomar exactamente frame_samples_8k muestras
                        if len(chunk) > frame_samples_8k:
                            chunk = chunk[:frame_samples_8k]
                        elif len(chunk) < frame_samples_8k // 2:
                            continue

                        energy   = float(np.mean(chunk ** 2))
                        ring_e   = self._goertzel_ring_energy(chunk, TARGET_SR)
                        flatness = self._spectral_flatness(chunk)
                        zcr      = self._zcr(chunk)

                        cls = self._classify_frame(energy, ring_e, flatness, zcr)
                        now = time.time()

                        self._update_ring_state(cls, now)
                        self._update_vad(cls)            # Método 1: espectral
                        self._update_energy_vad(energy)  # Método 2: energía sostenida

                        # ── Alimentar grabador con audio de DIALING ─────────────
                        if _active_recorder is not None:
                            raw_ch = data[:, 0].astype(np.float32) if data.ndim > 1 else data.astype(np.float32)
                            if cls == "ring":
                                n  = len(raw_ch)
                                t  = np.arange(n, dtype=np.float32) / sr_native
                                synth_ring = (0.45 * np.sin(2.0 * np.pi * 425.0 * t)).astype(np.float32)
                                _active_recorder.feed(synth_ring, sr_native)
                            else:
                                _active_recorder.feed(raw_ch, sr_native)

                        # ── Visualizador: emitir RMS al canal 'input' (~10 Hz) ──
                        rms = float(np.sqrt(energy))
                        socketio.emit("audio_viz", {"ch": "input", "rms": rms})

                break   # stream cerrado limpiamente — salir del retry loop

            except Exception as exc:
                last_exc = exc
                if self._stop_ev.is_set():
                    break
                print(f"[PreCall] Intento {attempt}/{max_retries} fallido: {exc}")
                if attempt < max_retries:
                    time.sleep(retry_delay)
                    retry_delay *= 1.5
                else:
                    print(f"[PreCall] ❌ Sin stream tras {max_retries} intentos: {last_exc}")

        print(f"[PreCall] Detenido — rings={self.ring_count} voice={self.operator_voice}")


class IVRCampaign(threading.Thread):
    """
    Hilo que procesa una cola de números telefónicos uno por uno.
    Para cada número:
      1. Marca con ADB
      2. Monitorea el estado via logcat (CallMonitor)
      3. Si ACTIVE → reproduce audio inicial
      4. Espera tono DTMF del cliente via monitor de micrófono
      5. Si detecta dígito válido → reproduce audio de despedida → cuelga
      6. Si no detecta en timeout → reproduce audio intermedio → repite
      7. Guarda resultado en CSV
    """

    def __init__(self, config: dict):
        super().__init__(daemon=True, name="IVRCampaign")
        self.config        = config
        self.queue         = deque(config.get("numbers", []))
        self.total         = len(self.queue)
        self.processed     = 0
        self.device_id     = config.get("device_id")
        self.delay_s       = float(config.get("delay_seconds",  IVR_DEFAULT_DELAY_S))
        self.audio_welcome = config.get("audio_welcome")
        self.audio_menu    = config.get("audio_menu")
        self.audio_bye     = config.get("audio_bye")
        self.audio_no_tone = config.get("audio_no_tone")
        self.ivr_options   = config.get("ivr_options", {})
        self.tone_timeout  = float(config.get("tone_timeout",  IVR_DEFAULT_TONE_TIMEOUT))
        self.menu_repeats  = int(config.get("menu_repeats",   IVR_DEFAULT_MENU_REPEATS))
        self.record_calls  = bool(config.get("record_calls",  False))
        self.is_test       = config.get("is_test", False)

        self._stop_event   = threading.Event()
        self._pause_event  = threading.Event()
        self._pause_event.set()   # no pausado al inicio

        # Para sincronizar detección de tono dentro de handle_active
        self._digit_event  = threading.Event()
        self._last_digit   = None

    # ── API pública ──────────────────────────────────────────────

    def stop(self):
        self._stop_event.set()
        self._digit_event.set()   # desbloquear espera de tono

    def pause(self):
        self._pause_event.clear()

    def resume(self):
        self._pause_event.set()

    @property
    def is_running(self):
        return self.is_alive() and not self._stop_event.is_set()

    def on_dtmf(self, digit: str):
        """Llamado por el handler WebSocket cuando llega un dígito DTMF."""
        print(f"[IVRCampaign] Digito recibido en campaña: {digit}")
        self._last_digit = digit
        self._digit_event.set()

    # ── Hilo principal ───────────────────────────────────────────

    def run(self):
        global _ivr_dtmf_callback
        _ivr_dtmf_callback = self.on_dtmf

        _emit_ivr("ivr_log", {"msg": f"🚀 Campaña iniciada — {self.total} números en cola",
                               "level": "info"})

        while self.queue and not self._stop_event.is_set():
            self._pause_event.wait()   # bloqueante si está pausado
            if self._stop_event.is_set():
                break

            number = self.queue.popleft()
            self.processed += 1
            self._process_number(number)

            if self.queue and not self._stop_event.is_set():
                _emit_ivr("ivr_log", {"msg": f"⏳ Esperando {self.delay_s}s antes de la próxima llamada",
                                       "level": "info"})
                self._interruptible_sleep(self.delay_s)

        _ivr_dtmf_callback = None
        status = "detenida" if self._stop_event.is_set() else "completada"
        _emit_ivr("ivr_log", {"msg": f"✅ Campaña {status}. Procesados: {self.processed}/{self.total}",
                               "level": "success"})
        _emit_ivr("ivr_campaign_done", {"processed": self.processed, "total": self.total})

    def _process_number(self, number: str):
        """Ejecuta el flujo completo para un número: marcar → monitorear → reaccionar."""
        global _active_recorder
        _emit_ivr("ivr_call_update", {
            "number": number, "status": "CALLING",
            "processed": self.processed, "total": self.total
        })

        # ── 3. Grabación + analizador pre-llamada (desde el primer momento) ─────
        # La grabación arranca al marcar para capturar ring tones, voz de operador
        # y toda la llamada desde el primer segundo.
        recorder = None
        if self.record_calls and _SD_OK and _audio_monitor_device is not None:
            ts       = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            safe_num = number.replace("+", "").replace(" ", "")
            rec_path = os.path.join(IVR_RECORDINGS_DIR, f"{ts}_{safe_num}.wav")
            recorder = CallRecorder(filepath=rec_path)
            _active_recorder = recorder
            recorder.start()
            _emit_ivr("ivr_log", {"msg": "  🔴 Grabando desde marcado [estéreo mic+IVR]...", "level": "info"})

        pre_call = None
        if _SD_OK and _audio_monitor_device is not None:
            # Detener el visualizador idle (_input_viz) si está usando el mismo
            # dispositivo de entrada — Windows WASAPI no permite dos streams al mismo tiempo.
            if _input_viz and _input_viz.is_alive():
                print("[IVR] Parando InputVizMonitor antes de PreCallAnalyzer...")
                _stop_all_input_monitors(join=True)

            pre_call = PreCallAudioAnalyzer(device_index=_audio_monitor_device)
            pre_call.start()
            _emit_ivr("ivr_log", {
                "msg": "  🔍 Analizando audio pre-llamada (ring/voz)...",
                "level": "info"
            })


        _emit_ivr("ivr_log", {"msg": f"📞 Marcando: {number}", "level": "info"})

        result_status = "NO_ANSWER"
        result_digit  = None

        # — 1. Marcar ——————————————————————————————————————————
        try:
            self._adb(["shell", "am", "start", "-a",
                       "android.intent.action.CALL", "-d", f"tel:{number}"])
        except Exception as exc:
            _emit_ivr("ivr_log", {"msg": f"❌ Error marcando {number}: {exc}", "level": "error"})
            _save_call_result(number, "ADB_ERROR", None, str(exc))
            _emit_ivr("ivr_call_update", {"number": number, "status": "ERROR",
                                           "processed": self.processed, "total": self.total})
            return

        # — 2. Monitorear estado —————————————————————————————
        # call_stop:        desbloquea el bucle de espera inicial (ACTIVE o DISCONNECTED temprano)
        # monitor_stop:     detiene el hilo logcat — SOLO al final del flujo completo
        # call_disconnected: se activa si el cliente cuelga EN CUALQUIER momento
        call_stop         = threading.Event()
        monitor_stop      = threading.Event()   # ← evento INDEPENDIENTE para el CallMonitor
        call_disconnected = threading.Event()
        call_state        = {"current": "CONNECTING"}
        # Timestamp en que la llamada entró a DIALING — usado para el discriminador de tiempo
        # Un número apagado/no disponible pasa a ACTIVE en < 10s desde DIALING
        # Un número que timbra (incluso al buzón) pasa a ACTIVE en ≥10s
        dialing_start_time: list = [time.time()]   # lista mutable para acceso desde closure
        MIN_DIALING_SECS = 15.0   # seg — umbral de tiempo en DIALING para confirmar que timbó

        def on_state(state: str):
            call_state["current"] = state
            _emit_ivr("ivr_status", {"number": number, "state": state})
            _emit_ivr("ivr_log", {"msg": f"  -> Estado: {state}", "level": "info"})
            if state == "DIALING":
                dialing_start_time[0] = time.time()   # registrar inicio de DIALING
            elif state == "ACTIVE":
                call_stop.set()          # desbloquea el bucle de espera inicial — NO detiene el monitor
            elif state == "DISCONNECTED":
                call_stop.set()          # desbloquea el bucle de espera inicial
                call_disconnected.set()  # señala cuelgue a _handle_active()
                # NO hacemos monitor_stop.set() aquí; el monitor se para al final del flujo

        if _MONITOR_OK:
            monitor = CallMonitor(device_id=self.device_id)
            # Usamos monitor_stop (no call_stop) para que el monitor siga leyendo
            # logcat incluso después de entrar a ACTIVE — así detecta el DISCONNECTED
            # que ocurre cuando el cliente cuelga durante la reproducción de audio.
            monitor.start(on_state_change=on_state, stop_event=monitor_stop, clear_logs=True)
        else:
            time.sleep(3)
            call_state["current"] = "ACTIVE"

        # ─ 3. Analizador de audio pre-llamada (durante DIALING) ──────
        # Detecta ring tones y voz del operador ANTES de que la llamada entre en ACTIVE.
        # El recorder ya está corriendo — PreCallAnalyzer también alimenta feed() si hay recorder.
        if _SD_OK and _audio_monitor_device is not None and pre_call is None:
            pre_call = PreCallAudioAnalyzer(device_index=_audio_monitor_device)
            pre_call.start()
            _emit_ivr("ivr_log", {
                "msg": "  🔍 Analizando audio pre-llamada (ring/voz)...",
                "level": "info"
            })

        # Esperar hasta que la llamada entre en ACTIVE o se desconecte (máx 60 s)
        deadline = time.time() + 60
        while not call_stop.is_set() and not self._stop_event.is_set():
            if time.time() > deadline:
                break
            time.sleep(0.2)

        # ── Ventana post-ACTIVE para detectar voz del operador ─────────────────
        # El audio del teléfono NO se enruta al PC durante DIALING en muchos dispositivos;
        # solo llega cuando Android reporta ACTIVE. Extendemos la captura del
        # PreCallAudioAnalyzer hasta POST_ACTIVE_LISTEN seg adicionales post-ACTIVE
        # (con salida anticipada en cuanto detecta voz sostenida).
        # Solo aplica cuando el tiempo en DIALING < MIN_DIALING_SECS (candidato a UNAVAILABLE).
        POST_ACTIVE_LISTEN = 4.5   # seg máx post-ACTIVE escuchando operador
        time_in_dialing = time.time() - dialing_start_time[0]

        if (pre_call and pre_call.is_alive()
                and call_state["current"] == "ACTIVE"
                and time_in_dialing < MIN_DIALING_SECS
                and not pre_call.operator_voice):
            # Reiniciar contador VAD: medimos VOZ fresca desde el inicio de ACTIVE
            # (descartamos lo que haya podido acumularse durante DIALING que era silencio)
            pre_call.reset_vad_counter()
            _emit_ivr("ivr_log", {
                "msg": f"  🎧 Escuchando {POST_ACTIVE_LISTEN}s post-ACTIVE para detectar voz de operador...",
                "level": "info"
            })
            post_deadline = time.time() + POST_ACTIVE_LISTEN
            while (time.time() < post_deadline
                   and not pre_call.operator_voice
                   and not self._stop_event.is_set()):
                time.sleep(0.1)
            if pre_call.operator_voice:
                _emit_ivr("ivr_log", {"msg": "  🔔 Voz de operador detectada post-ACTIVE", "level": "warn"})

        # Detener analizador pre-llamada y leer resultados
        pre_rings = 0
        pre_voice = False
        if pre_call and pre_call.is_alive():
            pre_call.stop()
            pre_call.join(timeout=2.0)
        if pre_call:
            pre_rings = pre_call.ring_count
            pre_voice = pre_call.operator_voice
            _emit_ivr("ivr_log", {
                "msg": f"  📊 Pre-llamada: {pre_rings} ring(s) — voz_operador={pre_voice}",
                "level": "info"
            })

        final_state = call_state["current"]
        # Nota: time_in_dialing ya fue calculado antes de la ventana post-ACTIVE
        # y refleja el tiempo real en DIALING (no incluye el tiempo de la ventana de escucha).

        if final_state == "ACTIVE":
            _emit_ivr("ivr_log", {
                "msg": f"  ⏱️ Tiempo en DIALING: {time_in_dialing:.1f}s",
                "level": "info"
            })
            # Condición UNAVAILABLE:
            #   1. El analizador detectó voz del operador  Y
            #   2. Se contaron pocos rings (indicador de audio)  Y
            #   3. El tiempo en DIALING fue < MIN_DIALING_SECS  ← GUARDIA PRINCIPAL
            #      (si estuvo ≥10s en DIALING, definitivamente timbó → NO es UNAVAILABLE)
            is_unavailable = (
                pre_voice
                and pre_rings <= PreCallAudioAnalyzer.MAX_RINGS
                and time_in_dialing < MIN_DIALING_SECS
            )
            if is_unavailable:
                # ⛔ Operador respondió sin dar tiempo a suficientes rings → apagado/sin servicio
                result_status = "UNAVAILABLE"
                _emit_ivr("ivr_log", {
                    "msg": f"  ⛔ Número no disponible — {pre_rings} ring(s), {time_in_dialing:.1f}s antes del operador",
                    "level": "warn"
                })
                self._hang_up()
            else:
                # El monitor sigue corriendo → puede detectar DISCONNECTED durante _handle_active()
                result_status, result_digit = self._handle_active(number, call_disconnected, recorder)
        elif final_state in ("DISCONNECTED", "CONNECTING", "DIALING"):
            result_status = "NO_ANSWER" if final_state != "DISCONNECTED" else "DISCONNECTED"
        else:
            result_status = "UNKNOWN"

        # Ahora sí detenemos el monitor logcat
        if _MONITOR_OK:
            monitor_stop.set()
            monitor.stop()

        self._hang_up()

        # ── Detener y guardar la grabación (todos los casos) ───────────────
        # El recorder inició al marcar → debemos detenerlo siempre, ya sea ACTIVE,
        # UNAVAILABLE, NO_ANSWER o cualquier otro resultado.
        if recorder:
            _active_recorder = None          # desconectar del monitor ANTES de stop
            recorder.stop()
            recorder.join(timeout=5.0)
            if os.path.isfile(recorder.filepath):
                kb = os.path.getsize(recorder.filepath) / 1024
                _emit_ivr("ivr_log", {
                    "msg": f"  💾 Grabación: {os.path.basename(recorder.filepath)} ({kb:.0f} KB) [{result_status}]",
                    "level": "success"
                })

        # Incluir info de pre-llamada en las notas del CSV cuando aplica
        pre_notes = ""
        if pre_call and (pre_voice or pre_rings > 0):
            pre_notes = f"rings={pre_rings} voice={pre_voice}"
        _save_call_result(number, result_status, result_digit, pre_notes)
        _emit_ivr("ivr_call_update", {
            "number": number, "status": result_status,
            "digit": result_digit,
            "processed": self.processed, "total": self.total
        })
        _emit_ivr("ivr_log", {
            "msg": f"  OK {number} -> {result_status}" + (f" (opcion: {result_digit})" if result_digit else ""),
            "level": "success"
        })

        # ── WhatsApp notification ────────────────────────────────────
        if _wa_config.get("enabled") and _wa_notifier:
            option_desc = None
            if result_digit and self.ivr_options:
                opt = self.ivr_options.get(result_digit)
                if isinstance(opt, dict):
                    option_desc = opt.get("desc", result_digit)
                elif opt:
                    option_desc = str(opt)
            _send_whatsapp_notification(number, result_status, result_digit, option_desc)

    def _handle_active(self, number: str,
                        disconnect_event: threading.Event = None,
                        recorder=None) -> tuple[str, str | None]:
        """
        La llamada fue contestada (ACTIVE).
        Reproduce audio inicial, espera tono válido, actúa.
        Tonos no configurados se ignoran y se sigue esperando.

        disconnect_event: evento que se activa si el cliente cuelga en cualquier momento.
                          Cuando se detecta, se interrumpe la reproducción de audio y
                          el bucle de espera de tono inmediatamente.

        Retorna (status, digit_detectado)
        """
        # Evento combinado: cuelgue del cliente O stop de la campaña
        def _caller_gone() -> bool:
            """True si el cliente colgó o se detuvo la campaña."""
            return self._stop_event.is_set() or (
                disconnect_event is not None and disconnect_event.is_set()
            )

        _emit_ivr("ivr_log", {"msg": "  Llamada contestada — activando monitor de audio...", "level": "success"})
        start_audio_monitor(_audio_monitor_device)

        self._digit_event.clear()
        self._last_digit = None

        # El grabador ya está corriendo desde _process_number (inicio de la llamada).
        # _stop_recorder aquí solo desconecta del monitor de audio (evita que le lleguen
        # más chunks). El stop/join real lo hace _process_number al terminar.
        def _stop_recorder():
            """Desconecta el grabador del monitor de audio (sin detener el thread)."""
            global _active_recorder
            if recorder and _active_recorder is recorder:
                _active_recorder = None   # el monitor ya no envía chunks al recorder

        try:
            # ══ 1. BIENVENIDA — se reproduce UNA sola vez ════════════════
            _emit_ivr("ivr_log", {"msg": "  🎙️ Reproduciendo bienvenida...", "level": "info"})
            _play_audio(self.audio_welcome, cancel_event=disconnect_event)

            if _caller_gone():
                stop_audio_monitor()
                if disconnect_event and disconnect_event.is_set():
                    # Verificar si el cliente ya pulsó una opción válida ANTES de colgar
                    saved = self._last_digit
                    if saved and saved in self.ivr_options:
                        _emit_ivr("ivr_log", {"msg": f"  ✅ Tono {saved!r} recibido antes del cuelgue — registrando respuesta", "level": "success"})
                        return "ANSWERED_TONE", saved
                    _emit_ivr("ivr_log", {"msg": "  📵 Cliente colgó durante la bienvenida", "level": "warn"})
                    return "DISCONNECTED_DURING_CALL", None
                return "STOPPED", None

            # ══ 2. MENÚ IVR — se repite menu_repeats veces ══════════════
            if self._digit_event.is_set():
                _emit_ivr("ivr_log", {"msg": "  (Señal detectada durante la bienvenida — evaluando...)", "level": "info"})

            for attempt in range(self.menu_repeats):
                if _caller_gone():
                    stop_audio_monitor()
                    if disconnect_event and disconnect_event.is_set():
                        # Verificar si el cliente ya pulsó una opción válida ANTES de colgar
                        saved = self._last_digit
                        if saved and saved in self.ivr_options:
                            _emit_ivr("ivr_log", {"msg": f"  ✅ Tono {saved!r} recibido antes del cuelgue — registrando respuesta", "level": "success"})
                            return "ANSWERED_TONE", saved
                        _emit_ivr("ivr_log", {"msg": "  📵 Cliente colgó antes del menú IVR", "level": "warn"})
                        return "DISCONNECTED_DURING_CALL", None
                    return "STOPPED", None

                _emit_ivr("ivr_log", {
                    "msg": f"  📋 Reproduciendo menú IVR (intento {attempt + 1}/{self.menu_repeats})...",
                    "level": "info"
                })
                _play_audio(self.audio_menu, cancel_event=disconnect_event)

                if _caller_gone():
                    stop_audio_monitor()
                    if disconnect_event and disconnect_event.is_set():
                        # Verificar si el cliente ya pulsó una opción válida ANTES de colgar
                        saved = self._last_digit
                        if saved and saved in self.ivr_options:
                            _emit_ivr("ivr_log", {"msg": f"  ✅ Tono {saved!r} recibido antes del cuelgue — registrando respuesta", "level": "success"})
                            return "ANSWERED_TONE", saved
                        _emit_ivr("ivr_log", {"msg": "  📵 Cliente colgó durante el menú IVR", "level": "warn"})
                        return "DISCONNECTED_DURING_CALL", None
                    return "STOPPED", None

                deadline = time.time() + self.tone_timeout
                _emit_ivr("ivr_log", {"msg": f"  ⏳ Esperando tono válido ({self.tone_timeout}s)...", "level": "info"})

                digit = None
                while time.time() < deadline:
                    if _caller_gone():
                        stop_audio_monitor()
                        if disconnect_event and disconnect_event.is_set():
                            # Prioridad: si ya había un dígito válido en este intento
                            # o en _last_digit, registrarlo como respuesta.
                            if digit and digit in self.ivr_options:
                                _emit_ivr("ivr_log", {"msg": f"  ✅ Tono {digit!r} ya validado — registrando respuesta (cuelgue ignorado)", "level": "success"})
                                return "ANSWERED_TONE", digit
                            saved = self._last_digit
                            if saved and saved in self.ivr_options:
                                _emit_ivr("ivr_log", {"msg": f"  ✅ Tono {saved!r} recibido antes del cuelgue — registrando respuesta", "level": "success"})
                                return "ANSWERED_TONE", saved
                            _emit_ivr("ivr_log", {"msg": "  📵 Cliente colgó mientras esperaba tono", "level": "warn"})
                            return "DISCONNECTED_DURING_CALL", None
                        return "STOPPED", None

                    remaining = max(0.1, deadline - time.time())
                    self._digit_event.wait(timeout=remaining)
                    self._digit_event.clear()
                    candidate = self._last_digit
                    self._last_digit = None
                    if not candidate:
                        continue
                    if candidate in self.ivr_options:
                        digit = candidate
                        break
                    else:
                        _emit_ivr("ivr_log", {
                            "msg": f"  Tono '{candidate}' no configurado — ignorando",
                            "level": "info"
                        })

                if digit:
                    option_data = self.ivr_options[digit]
                    if isinstance(option_data, dict):
                        desc      = option_data.get("desc", digit)
                        audio_bye = option_data.get("audio_bye") or self.audio_bye
                    else:
                        desc      = str(option_data)
                        audio_bye = self.audio_bye

                    _emit_ivr("ivr_log", {"msg": f"  ✅ Tono válido: {digit} — {desc}", "level": "success"})
                    _emit_ivr("ivr_digit", {"number": number, "digit": digit, "option": desc})
                    _play_audio(audio_bye, cancel_event=disconnect_event)
                    self._hang_up()
                    stop_audio_monitor()
                    return "ANSWERED_TONE", digit

            # ══ 3. SIN TONO — agotados todos los intentos ════════════════
            _emit_ivr("ivr_log", {"msg": "  ⚠️ Sin tono detectado en todos los intentos", "level": "warn"})
            if self.audio_no_tone:
                _emit_ivr("ivr_log", {"msg": "  🔔 Reproduciendo audio de cierre (sin tono)...", "level": "info"})
                _play_audio(self.audio_no_tone, cancel_event=disconnect_event)
            self._hang_up()
            stop_audio_monitor()
            return "ANSWERED_NO_TONE", None

        finally:
            _stop_recorder()   # siempre se ejecuta — en cualquier punto de salida

    # ── Helpers ADB ──────────────────────────────────────────────

    def _adb(self, args: list):
        cmd = ["adb"]
        if self.device_id:
            cmd += ["-s", self.device_id]
        cmd += args
        return subprocess.run(cmd, capture_output=True, timeout=10)

    def _hang_up(self):
        try:
            self._adb(["shell", "input", "keyevent", "6"])  # KEYCODE_ENDCALL
        except Exception:
            pass

    def _interruptible_sleep(self, seconds: float):
        deadline = time.time() + seconds
        while time.time() < deadline and not self._stop_event.is_set():
            time.sleep(0.2)


# ══════════════════════════════════════════════
#  MONITOR DE AUDIO EN PYTHON (sounddevice)
# ══════════════════════════════════════════════

_audio_monitor_thread: "PythonAudioMonitor | None" = None
_audio_monitor_device = None   # índice del dispositivo seleccionado

class PythonAudioMonitor(threading.Thread):
    """
    Captura audio directamente del dispositivo de entrada del sistema
    y ejecuta detección DTMF Goertzel en tiempo real.
    No depende del navegador ni de mediaDevices.

    Para multi-sesión: pasar `dtmf_callback` al constructor para aislar
    el enrutamiento DTMF de esta instancia del global _ivr_dtmf_callback.
    """
    WINDOW = int(TARGET_SR * 0.08)   # 640 muestras @ 8 kHz
    HOP    = int(TARGET_SR * 0.02)   # 160 muestras

    def __init__(self, device_index=None, dtmf_callback=None,
                 session_id: str = None):
        super().__init__(daemon=True, name="PythonAudioMonitor")
        self.device_index    = device_index
        # Callback DTMF por instancia (multi-sesión) — si None usa global
        self._dtmf_callback  = dtmf_callback
        self._session_id     = session_id   # para filtrar eventos Socket.IO
        self._stop_ev = threading.Event()
        self._buf: list[float] = []
        self._last_digit: str | None = None
        self._last_emit = 0.0
        self._viz_acc   = 0.0
        self._viz_n     = 0

    def _emit_log(self, msg: str, level: str = "info"):
        """Emite log al canal correcto (global o sesión)."""
        if self._session_id:
            socketio.emit("session_log", {
                "session_id": self._session_id, "msg": msg, "level": level,
                "ts": time.time()
            })
        else:
            _emit_ivr("ivr_log", {"msg": msg, "level": level})

    def run(self):
        if not _SD_OK:
            self._emit_log("❌ sounddevice no instalado", "error")
            return
        try:
            dev_info = sd.query_devices(self.device_index, "input")
            sr_native = int(dev_info["default_samplerate"])
            dev_name  = dev_info["name"]
        except Exception as exc:
            self._emit_log(f"❌ Dispositivo audio inválido: {exc}", "error")
            return

        self._emit_log(f"🎤 Monitor activo: [{dev_name}] @ {sr_native} Hz", "success")

        def callback(indata, frames, time_info, status):
            if self._stop_ev.is_set():
                raise sd.CallbackStop()

            chunk = indata[:, 0].astype(np.float32)

            # ── Visualizador: emitir RMS nativo ~15 Hz ────────────────────
            rms_raw = float(np.sqrt(np.mean(chunk ** 2)))
            self._viz_acc += rms_raw
            self._viz_n   += 1
            if self._viz_n >= 3:   # ~15 Hz si blocksize=40ms
                if self._session_id:
                    socketio.emit("session_viz", {
                        "session_id": self._session_id,
                        "ch": "input", "rms": self._viz_acc / self._viz_n
                    })
                else:
                    socketio.emit("audio_viz", {"ch": "input", "rms": self._viz_acc / self._viz_n})
                self._viz_acc = 0.0
                self._viz_n   = 0

            # ── Alimentar grabador: sintetizar DTMF si hay dígito activo ──────
            if _active_recorder is not None:
                if self._last_digit and self._last_digit in _DTMF_DIGIT_FREQS:
                    f_row, f_col = _DTMF_DIGIT_FREQS[self._last_digit]
                    n_s = len(chunk)
                    t_s = np.arange(n_s, dtype=np.float32) / sr_native
                    chunk_rec = (
                        0.35 * np.sin(2.0 * np.pi * f_row * t_s) +
                        0.35 * np.sin(2.0 * np.pi * f_col * t_s)
                    ).astype(np.float32)
                else:
                    chunk_rec = chunk
                _active_recorder.feed(chunk_rec, sr_native)

            # Resamplear a 8 kHz para análisis DTMF
            if sr_native != TARGET_SR:
                g = gcd(TARGET_SR, sr_native)
                chunk = resample_poly(chunk, TARGET_SR // g, sr_native // g).astype(np.float32)

            self._buf.extend(chunk.tolist())
            if len(self._buf) < self.WINDOW:
                return

            frame  = np.array(self._buf[-self.WINDOW:], dtype=np.float32)
            sos    = _get_bandpass(TARGET_SR)
            frame  = sosfilt(sos, frame).astype(np.float32)
            energy = float(np.mean(frame ** 2))

            if energy < ENERGY_THRESHOLD:
                self._last_digit = None
                if len(self._buf) > self.WINDOW * 4:
                    self._buf = self._buf[-self.WINDOW:]
                return

            digit = detect_dtmf_frame(frame, TARGET_SR, energy)

            # Emitir estado a la UI
            now = time.time()
            if now - self._last_emit > 0.08:
                self._last_emit = now
                socketio.emit("rt_digit", {"digit": digit, "energy": round(float(energy), 8),
                                           "session_id": self._session_id})

            # Notificar al IVR cuando hay dígito nuevo
            if digit and digit != self._last_digit:
                self._last_digit = digit
                print(f"[AudioMonitor] Dígito: {digit}  E={energy:.2e}"
                      + (f"  [s:{self._session_id}]" if self._session_id else ""))
                self._emit_log(f"  🎯 Tono detectado: {digit}", "success")

                # Usar callback por instancia (multi-sesión) o global (legacy)
                cb = self._dtmf_callback or _ivr_dtmf_callback
                if cb:
                    try:
                        cb(digit)
                    except Exception as exc:
                        print(f"[AudioMonitor] Error callback: {exc}")
            elif not digit:
                self._last_digit = None

            if len(self._buf) > self.WINDOW * 4:
                self._buf = self._buf[-self.WINDOW:]

        # ── Abrir stream con retry (WASAPI puede tardar en liberar el device) ──
        max_retries = 4
        retry_delay = 0.3
        last_exc    = None

        for attempt in range(1, max_retries + 1):
            try:
                with sd.InputStream(
                    device    = self.device_index,
                    channels  = 1,
                    samplerate= sr_native,
                    blocksize = int(sr_native * 0.04),   # 40ms
                    dtype     = "float32",
                    callback  = callback,
                ):
                    if attempt > 1:
                        self._emit_log(f"🎤 Stream abierto (intento {attempt})", "success")
                    self._stop_ev.wait()
                break

            except Exception as exc:
                last_exc = exc
                if self._stop_ev.is_set():
                    break
                print(f"[AudioMonitor] Intento {attempt}/{max_retries} fallido: {exc}")
                if attempt < max_retries:
                    time.sleep(retry_delay)
                    retry_delay *= 1.5
                else:
                    self._emit_log(
                        f"❌ Monitor audio: sin stream tras {max_retries} intentos — {last_exc}",
                        "error"
                    )

        self._emit_log("🔇 Monitor de audio detenido", "info")

    def stop(self):
        self._stop_ev.set()



# ══════════════════════════════════════════════
#  LOOPBACK ENERGY MONITOR — visualizador canal de SALIDA
# ══════════════════════════════════════════════

class LoopbackEnergyMonitor(threading.Thread):
    """
    Captura el audio reproducido por el sistema (WASAPI loopback) y emite
    eventos 'audio_viz' con ch='output' para el visualizador de la UI.
    """
    def __init__(self, device_index):
        super().__init__(daemon=True, name="LoopbackViz")
        self.device_index = device_index
        self._stop_ev     = threading.Event()

    def stop(self):
        self._stop_ev.set()

    def run(self):
        if not _SD_OK or self.device_index is None:
            return
        try:
            out_info = sd.query_devices(self.device_index, "output")
            sr_out   = int(out_info["default_samplerate"])
            bs       = int(sr_out * 0.04)   # bloques de 40 ms
            wasapi   = sd.WasapiSettings(loopback=True)
            acc, n   = 0.0, 0
            with sd.InputStream(
                device        = self.device_index,
                channels      = 1,
                samplerate    = sr_out,
                blocksize     = bs,
                dtype         = "float32",
                extra_settings= wasapi,
            ) as stream:
                while not self._stop_ev.is_set():
                    data, _ = stream.read(bs)
                    rms  = float(np.sqrt(np.mean(data ** 2)))
                    acc += rms
                    n   += 1
                    if n >= 3:   # ~15 Hz
                        socketio.emit("audio_viz", {"ch": "output", "rms": acc / n})
                        acc, n = 0.0, 0
        except Exception as exc:
            print(f"[LoopbackViz] {exc}")


_loopback_viz: LoopbackEnergyMonitor | None = None
_input_viz:    "InputVizMonitor | None" = None


class InputVizMonitor(threading.Thread):
    """
    Captura audio de la interfaz de entrada seleccionada y emite audio_viz
    para el visualizador de la UI. Opera independientemente de las llamadas.
    """
    def __init__(self, device_index):
        super().__init__(daemon=True, name="InputViz")
        self.device_index = device_index
        self._stop_ev     = threading.Event()

    def stop(self):
        self._stop_ev.set()

    def run(self):
        if not _SD_OK or self.device_index is None:
            return
        try:
            in_info = sd.query_devices(self.device_index, "input")
            sr      = int(in_info["default_samplerate"])
            bs      = int(sr * 0.04)   # bloques de 40 ms
            acc, n  = 0.0, 0
            with sd.InputStream(
                device    = self.device_index,
                channels  = 1,
                samplerate= sr,
                blocksize = bs,
                dtype     = "float32",
            ) as stream:
                while not self._stop_ev.is_set():
                    data, _ = stream.read(bs)
                    rms     = float(np.sqrt(np.mean(data ** 2)))
                    acc    += rms
                    n      += 1
                    if n >= 3:   # ~15 Hz
                        socketio.emit("audio_viz", {"ch": "input", "rms": acc / n})
                        acc, n = 0.0, 0
        except Exception as exc:
            print(f"[InputViz] {exc}")


# ── Tiempo de cortesia entre stop y start de un stream (ms) ───────────────────
# Windows WASAPI necesita ~150-200ms para liberar el dispositivo tras close().
# Sin este delay, el siguiente open() falla silenciosamente o usa el dispositivo
# default en lugar del configurado.
_STREAM_RELEASE_SLEEP = 0.25   # segundos
_STREAM_JOIN_TIMEOUT  = 5.0    # segundos (antes era 2.0, insuficiente en WASAPI)


def _stop_all_input_monitors(join: bool = True) -> None:
    """
    Detiene TODOS los monitores de entrada activos antes de abrir uno nuevo.
    Garantiza que el dispositivo de entrada quede libre en el OS.

    Monitores gestionados:
      _audio_monitor_thread  (PythonAudioMonitor  — DTMF durante llamada)
      _input_viz             (InputVizMonitor     — visualizador idle)

    El join tiene timeout _STREAM_JOIN_TIMEOUT para no bloquear indefinidamente.
    Tras el join se espera _STREAM_RELEASE_SLEEP para que el OS libere WASAPI.
    """
    global _audio_monitor_thread, _input_viz

    stopped_any = False

    if _audio_monitor_thread and _audio_monitor_thread.is_alive():
        _audio_monitor_thread.stop()
        if join:
            _audio_monitor_thread.join(timeout=_STREAM_JOIN_TIMEOUT)
        _audio_monitor_thread = None
        stopped_any = True

    if _input_viz and _input_viz.is_alive():
        _input_viz.stop()
        if join:
            _input_viz.join(timeout=_STREAM_JOIN_TIMEOUT)
        _input_viz = None
        stopped_any = True

    if stopped_any:
        # Dar tiempo al OS para liberar el handle WASAPI antes del siguiente open()
        time.sleep(_STREAM_RELEASE_SLEEP)


def _stop_all_output_monitors(join: bool = True) -> None:
    """
    Detiene TODOS los monitores de salida activos (loopback WASAPI).
    Necesario antes de abrir un nuevo LoopbackEnergyMonitor o reproducir con pygame.
    """
    global _loopback_viz

    if _loopback_viz and _loopback_viz.is_alive():
        _loopback_viz.stop()
        if join:
            _loopback_viz.join(timeout=_STREAM_JOIN_TIMEOUT)
        _loopback_viz = None
        time.sleep(_STREAM_RELEASE_SLEEP)


def start_audio_monitor(device_index=None):
    """
    Inicia el PythonAudioMonitor para captura DTMF durante la llamada.

    Antes de abrir el nuevo stream:
      1. Para _audio_monitor_thread (si existe)
      2. Para _input_viz (MISMO dispositivo — conflicto WASAPI)
      3. Espera _STREAM_RELEASE_SLEEP para que el OS libere el handle

    El LoopbackEnergyMonitor (salida) se reinicia solo si hay dispositivo configurado.
    """
    global _audio_monitor_thread, _loopback_viz

    # Detener TODOS los streams de entrada (evitar conflicto de dispositivo)
    _stop_all_input_monitors(join=True)

    # Iniciar PythonAudioMonitor (con retry incorporado en su run())
    _audio_monitor_thread = PythonAudioMonitor(device_index)
    _audio_monitor_thread.start()

    # Iniciar LoopbackEnergyMonitor (canal de salida) si hay dispositivo
    if _audio_output_device_index is not None:
        _stop_all_output_monitors(join=True)
        _loopback_viz = LoopbackEnergyMonitor(_audio_output_device_index)
        _loopback_viz.start()


def stop_audio_monitor():
    """Detiene monitor DTMF y visualizador loopback de salida."""
    _stop_all_input_monitors(join=False)   # no bloquear — la llamada ya terminó
    _stop_all_output_monitors(join=False)


@app.route("/ivr/viz/start", methods=["POST"])
def ivr_viz_start():
    """Inicia monitores de visualización de audio (entrada + salida) independientes de llamadas."""
    global _input_viz, _loopback_viz
    data       = request.get_json(force=True) or {}
    in_idx     = data.get("input")
    out_idx    = data.get("output")

    # ── Monitor de entrada ──────────────────────────────
    if in_idx is not None and _SD_OK:
        if _input_viz and _input_viz.is_alive():
            _input_viz.stop()
            _input_viz.join(timeout=1.5)
        _input_viz = InputVizMonitor(int(in_idx))
        _input_viz.start()
        print(f"[Viz] Monitor entrada iniciado — idx={in_idx}")

    # ── Monitor de salida (loopback WASAPI) ───────────────────
    if out_idx is not None and _SD_OK:
        if _loopback_viz and _loopback_viz.is_alive():
            _loopback_viz.stop()
            _loopback_viz.join(timeout=1.5)
        _loopback_viz = LoopbackEnergyMonitor(int(out_idx))
        _loopback_viz.start()
        print(f"[Viz] Monitor salida (loopback) iniciado — idx={out_idx}")

    return jsonify({"ok": True})


@app.route("/ivr/viz/stop", methods=["POST"])
def ivr_viz_stop():
    """
    Detiene los monitores de visualización de audio.
    Hace join real (con timeout) para que el OS libere el device
    antes de que el cliente vuelva a llamar a /ivr/viz/start.
    """
    _stop_all_input_monitors(join=True)
    _stop_all_output_monitors(join=True)
    return jsonify({"ok": True})


@app.route("/ivr/audio_devices")
def ivr_audio_devices():
    """Lista los dispositivos de audio de entrada Y salida del sistema."""
    if not _SD_OK:
        return jsonify({"ok": False, "error": "sounddevice no instalado",
                        "inputs": [], "outputs": []}), 200
    try:
        devices   = sd.query_devices()
        default_in  = sd.default.device[0]
        default_out = sd.default.device[1]

        # Coleccionar entradas y salidas sin duplicar por nombre
        seen_in, seen_out = set(), set()
        inputs, outputs   = [], []

        for i, d in enumerate(devices):
            name = d["name"]
            sr   = int(d["default_samplerate"])
            base = {"index": i, "name": name, "samplerate": sr}

            if d["max_input_channels"] >= 1 and name not in seen_in:
                seen_in.add(name)
                inputs.append({**base,
                               "channels":   d["max_input_channels"],
                               "is_default": (i == default_in)})

            if d["max_output_channels"] >= 1 and name not in seen_out:
                seen_out.add(name)
                outputs.append({**base,
                                "channels":   d["max_output_channels"],
                                "is_default": (i == default_out)})

        return jsonify({"ok": True, "inputs": inputs, "outputs": outputs})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc), "inputs": [], "outputs": []}), 200


@app.route("/ivr/monitor/start", methods=["POST"])
def ivr_monitor_start():
    """Inicia el monitor de audio en Python."""
    data = request.get_json(silent=True) or {}
    device_index = data.get("device_index")  # None = predeterminado
    global _audio_monitor_device
    _audio_monitor_device = device_index
    start_audio_monitor(device_index)
    return jsonify({"ok": True, "msg": "Monitor de audio iniciado"})


@app.route("/ivr/monitor/stop", methods=["POST"])
def ivr_monitor_stop():
    """Detiene el monitor de audio en Python."""
    stop_audio_monitor()
    return jsonify({"ok": True, "msg": "Monitor detenido"})


@app.route("/ivr/test_output", methods=["POST"])
def ivr_test_output():
    """Reproduce un pitido de prueba en el dispositivo de salida seleccionado."""
    if not _SD_OK:
        return jsonify({"ok": False, "error": "sounddevice no instalado"})
    data = request.get_json(silent=True) or {}
    device_index = data.get("device_index")  # None = predeterminado

    def _play_beep():
        global _loopback_viz
        # ── Pausar el loopback viz para liberar el dispositivo ──
        _lv_was_running = _loopback_viz and _loopback_viz.is_alive()
        if _lv_was_running:
            _loopback_viz.stop()
            _loopback_viz.join(timeout=1.5)
            _loopback_viz = None
        try:
            sr   = 44100
            dur  = 1.5      # 1.5 segundos
            freq = 1000.0   # 1 kHz — tono fácil de reconocer
            t    = np.linspace(0, dur, int(sr * dur), endpoint=False)
            env  = np.where(t < 0.05, t / 0.05,
                   np.where(t > 1.45, (dur - t) / 0.05, 1.0))
            wave = (np.sin(2 * np.pi * freq * t) * env * 0.7).astype(np.float32)

            # Emitir audio_viz durante la reproducción (simula nivel de salida)
            def _out_viz_thread():
                block = int(sr * 0.067)   # ~15 Hz
                for i in range(0, len(wave), block):
                    chunk = wave[i:i + block]
                    rms   = float(np.sqrt(np.mean(chunk ** 2)))
                    socketio.emit("audio_viz", {"ch": "output", "rms": rms})
                    time.sleep(0.067)

            threading.Thread(target=_out_viz_thread, daemon=True).start()

            kwargs = {"samplerate": sr}
            if device_index is not None:
                kwargs["device"] = int(device_index)
            sd.play(wave, **kwargs)
            sd.wait()
            _emit_ivr("ivr_log", {"msg": "🔊 Pitido de prueba reproducido OK", "level": "success"})
        except Exception as exc:
            _emit_ivr("ivr_log", {"msg": f"❌ Error reproduciendo pitido: {exc}", "level": "error"})
        finally:
            # ── Reiniciar loopback viz si estaba corriendo ──
            if _lv_was_running and _audio_output_device_index is not None:
                _loopback_viz = LoopbackEnergyMonitor(_audio_output_device_index)
                _loopback_viz.start()

    threading.Thread(target=_play_beep, daemon=True).start()
    return jsonify({"ok": True, "msg": "Reproduciendo pitido..."})


@app.route("/ivr/test_input", methods=["POST"])
def ivr_test_input():
    """
    Captura 3 segundos de audio del mic seleccionado y emite el nivel de energía
    por socket. El frontend muestra una barra de nivel en tiempo real Y la waveform.
    """
    if not _SD_OK:
        return jsonify({"ok": False, "error": "sounddevice no instalado"})
    data = request.get_json(silent=True) or {}
    device_index = data.get("device_index")

    def _capture():
        global _input_viz
        # ── Pausar InputVizMonitor para liberar el dispositivo ──
        _iv_was_running = _input_viz and _input_viz.is_alive()
        if _iv_was_running:
            _input_viz.stop()
            _input_viz.join(timeout=1.5)
            _input_viz = None
        try:
            dev_info  = sd.query_devices(device_index, "input")
            sr_native = int(dev_info["default_samplerate"])
            dev_name  = dev_info["name"]
            _emit_ivr("ivr_log", {"msg": f"🎤 Probando entrada: [{dev_name}]...", "level": "info"})

            duration = 3.0
            frames   = []
            stop_ev  = threading.Event()

            def callback(indata, n, t, status):
                chunk  = indata[:, 0].astype(np.float32)
                energy = float(np.mean(chunk ** 2))
                rms    = float(np.sqrt(energy))
                # Barra de nivel (0-100)
                level = 0 if energy < 1e-10 else min(100, int(10 * np.log10(energy / 1e-10)))
                socketio.emit("input_test_level", {"level": level, "energy": round(energy, 8)})
                # Waveform viz
                socketio.emit("audio_viz", {"ch": "input", "rms": rms})
                frames.append(chunk)

            with sd.InputStream(device=device_index, channels=1,
                                samplerate=sr_native, blocksize=int(sr_native * 0.067),
                                callback=callback):
                stop_ev.wait(timeout=duration)

            peak = max((float(np.max(np.abs(np.concatenate(frames)))) if frames else 0), 0)
            _emit_ivr("ivr_log", {
                "msg": f"✅ Prueba completada. Pico: {peak:.4f} {'(OK — se recibe señal)' if peak > 0.001 else '(SILENCIO — verifica el mic)'}",
                "level": "success" if peak > 0.001 else "warn"
            })
            socketio.emit("input_test_done", {"peak": round(peak, 5)})
        except Exception as exc:
            _emit_ivr("ivr_log", {"msg": f"❌ Error prueba entrada: {exc}", "level": "error"})
            socketio.emit("input_test_done", {"peak": 0})
        finally:
            # ── Reiniciar InputVizMonitor si estaba corriendo ──
            if _iv_was_running and device_index is not None:
                _input_viz = InputVizMonitor(int(device_index))
                _input_viz.start()

    threading.Thread(target=_capture, daemon=True).start()
    return jsonify({"ok": True, "msg": "Capturando 3 segundos..."})


# ══════════════════════════════════════════════
#  IVR AUTOMATOR — Rutas Flask
# ══════════════════════════════════════════════

@app.route("/ivr/devices")
def ivr_devices():
    """Lista los dispositivos ADB conectados."""
    try:
        # Intentar encontrar adb en el PATH o con where
        adb_cmd = "adb"
        try:
            check = subprocess.run(["adb", "version"], capture_output=True, timeout=3)
            if check.returncode != 0:
                raise FileNotFoundError
        except (FileNotFoundError, OSError):
            # En Windows intentar con 'where'
            where = subprocess.run(["where", "adb"], capture_output=True, text=True, timeout=3)
            if where.returncode == 0 and where.stdout.strip():
                adb_cmd = where.stdout.strip().splitlines()[0]
            else:
                return jsonify({"ok": False, "error": "adb no encontrado en el sistema. Instala Android SDK Platform-Tools y agrégalo al PATH."}), 500

        result = subprocess.run([adb_cmd, "devices"], capture_output=True, text=True, timeout=5)
        lines = result.stdout.strip().splitlines()
        devices = []
        for line in lines[1:]:  # saltar cabecera
            line = line.strip()
            if line and "\t" in line:
                serial, state = line.split("\t", 1)
                if state.strip() == "device":
                    devices.append(serial.strip())
        return jsonify({"ok": True, "devices": devices})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/ivr/upload_numbers", methods=["POST"])
def ivr_upload_numbers():
    """Recibe un Excel, extrae la columna 'Celular' y devuelve la lista."""
    if not _OPENPYXL_OK:
        return jsonify({"ok": False, "error": "openpyxl no instalado"}), 500

    f = request.files.get("file")
    if not f:
        return jsonify({"ok": False, "error": "No se recibió archivo"}), 400

    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb.active

        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col_idx = None
        for i, h in enumerate(headers):
            if h.lower() == "celular":
                col_idx = i
                break

        if col_idx is None:
            return jsonify({"ok": False, "error": "No se encontró la columna 'Celular'"}), 400

        numbers = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None:
                num = str(val).strip().replace(" ", "").replace("-", "")
                if num:
                    numbers.append(num)

        wb.close()
        return jsonify({"ok": True, "numbers": numbers, "count": len(numbers)})
    except Exception as exc:
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/ivr/upload_audio", methods=["POST"])
def ivr_upload_audio():
    """Guarda un archivo de audio para uso en el IVR. Tipo: 'initial'|'middle'|'bye'."""
    f = request.files.get("file")
    audio_type = request.form.get("type", "initial")  # initial | middle | bye
    if not f:
        return jsonify({"ok": False, "error": "No se recibió archivo"}), 400

    ext = os.path.splitext(f.filename)[1].lower()
    safe_name = f"ivr_{audio_type}{ext}"
    dest = os.path.join(IVR_AUDIO_FOLDER, safe_name)
    f.save(dest)
    return jsonify({"ok": True, "path": dest, "filename": f.filename, "type": audio_type})


@app.route("/ivr/wa/config", methods=["GET"])
def wa_config_get():
    """Devuelve la configuración actual de notificaciones WhatsApp."""
    return jsonify({
        "ok":      True,
        "config":  _wa_config,
        "available": _WA_OK,
        "browser": _wa_notifier.get_status() if _wa_notifier else {"status": "unavailable"},
    })


@app.route("/ivr/wa/config", methods=["POST"])
def wa_config_post():
    """Guarda la configuración de notificaciones WhatsApp."""
    global _wa_config
    data = request.get_json(force=True) or {}
    _wa_config["enabled"] = bool(data.get("enabled", False))
    _wa_config["contact"] = str(data.get("contact", "")).strip()
    _wa_config["backup"]  = str(data.get("backup",  "")).strip()
    _wa_save_config()
    return jsonify({"ok": True, "config": _wa_config})


@app.route("/ivr/wa/open_browser", methods=["POST"])
def wa_open_browser():
    """Abre Chrome con el perfil persistente de WhatsApp."""
    if not _WA_OK or not _wa_notifier:
        return jsonify({"ok": False, "error": "selenium no disponible"}), 500
    ok, msg = _wa_notifier.open_browser()
    return jsonify({"ok": ok, "msg": msg})


@app.route("/ivr/wa/close_browser", methods=["POST"])
def wa_close_browser():
    """Cierra el navegador de WhatsApp."""
    if not _wa_notifier:
        return jsonify({"ok": False, "error": "Notificador no disponible"}), 500
    ok, msg = _wa_notifier.close_browser()
    return jsonify({"ok": ok, "msg": msg})


@app.route("/ivr/wa/status")
def wa_status():
    """Estado actual del navegador de notificaciones."""
    if not _wa_notifier:
        return jsonify({"status": "unavailable", "message": "selenium no instalado", "available": False})
    s = _wa_notifier.get_status()
    s["available"] = _WA_OK
    return jsonify(s)


@app.route("/ivr/start", methods=["POST"])
def ivr_start():
    """Inicia una campaña o una prueba con 1 número."""
    global _ivr_campaign, _audio_monitor_device, _audio_output_device_name, _audio_output_device_index
    with _ivr_lock:
        if _ivr_campaign and _ivr_campaign.is_running:
            return jsonify({"ok": False, "error": "Ya hay una campaña activa"}), 409

    # ── Guard: no permitir campaña si hay llamada manual en curso ────────
    if _manual_call and _manual_call.is_active:
        return jsonify({
            "ok":    False,
            "error": "Hay una llamada manual en curso. Cuelga primero antes de iniciar una campaña."
        }), 409

    data = request.get_json(force=True) or {}

    numbers = data.get("numbers", [])
    if not numbers:
        return jsonify({"ok": False, "error": "Sin números en la cola"}), 400

    # Guardar dispositivos de audio seleccionados
    audio_input_index  = data.get("audio_device")        # índice int | None
    audio_output_index = data.get("audio_output_device") # índice int | None
    _audio_monitor_device       = audio_input_index
    _audio_output_device_index  = int(audio_output_index) if audio_output_index is not None else None

    # Obtener nombre del dispositivo de salida para pygame
    if audio_output_index is not None and _SD_OK:
        try:
            _audio_output_device_name = sd.query_devices(int(audio_output_index))["name"]
            print(f"[IVR] Salida de audio: {_audio_output_device_name}")
        except Exception:
            _audio_output_device_name = None
    else:
        _audio_output_device_name = None

    config = {
        "numbers":        numbers,
        "device_id":      data.get("device_id"),
        "delay_seconds":  data.get("delay_seconds", 5),
        "audio_welcome":  data.get("audio_welcome"),
        "audio_menu":     data.get("audio_menu"),
        "audio_bye":      data.get("audio_bye"),
        "audio_no_tone":  data.get("audio_no_tone"),
        "ivr_options":    data.get("ivr_options", {}),
        "tone_timeout":   data.get("tone_timeout", 10),
        "menu_repeats":   data.get("menu_repeats", 2),
        "record_calls":   bool(data.get("record_calls", False)),  # grabar llamadas contestadas
        "is_test":        data.get("is_test", False),
    }

    with _ivr_lock:
        _ivr_campaign = IVRCampaign(config)
        _ivr_campaign.start()

    # ── Iniciar watchdog ADB ─────────────────────────────────────
    global _adb_watchdog
    device_id = config.get("device_id")
    if device_id:
        if _adb_watchdog and _adb_watchdog.is_alive():
            _adb_watchdog.stop()
        _adb_watchdog = ADBWatchdog(device_id)
        _adb_watchdog.start()

    # ── Asegurar browser WA si notificaciones activas ────────────────
    if _wa_config.get("enabled") and _wa_notifier:
        contact = _wa_config.get("contact", "").strip()
        if contact:
            threading.Thread(
                target=_wa_notifier.ensure_ready,
                daemon=True,
                name="WA-EnsureReady"
            ).start()
            _emit_ivr("ivr_log", {"msg": "🔔 Verificando WhatsApp para notificaciones…", "level": "info"})

    mode = "prueba" if config["is_test"] else "campaña"
    return jsonify({"ok": True, "msg": f"{mode.capitalize()} iniciada", "total": len(numbers)})


@app.route("/ivr/stop", methods=["POST"])
def ivr_stop():
    """Detiene la campaña activa."""
    global _ivr_campaign, _adb_watchdog
    with _ivr_lock:
        if not _ivr_campaign or not _ivr_campaign.is_running:
            return jsonify({"ok": False, "error": "No hay campaña activa"}), 409
        _ivr_campaign.stop()
    # Detener watchdog ADB también
    if _adb_watchdog and _adb_watchdog.is_alive():
        _adb_watchdog.stop()
        _adb_watchdog = None
    return jsonify({"ok": True, "msg": "Campaña detenida"})


@app.route("/ivr/adb/status")
def ivr_adb_status():
    """Verifica en tiempo real si el dispositivo ADB sigue conectado."""
    device_id = request.args.get("device_id", "").strip()
    if not device_id:
        return jsonify({"connected": False, "error": "Sin device_id"})
    try:
        r = subprocess.run(
            ["adb", "-s", device_id, "get-state"],
            capture_output=True, text=True, timeout=5
        )
        connected = r.returncode == 0 and "device" in r.stdout
        return jsonify({"connected": connected, "device_id": device_id})
    except Exception as exc:
        return jsonify({"connected": False, "device_id": device_id, "error": str(exc)})



@app.route("/ivr/status")
def ivr_status():
    """Estado actual de la campaña."""
    global _ivr_campaign
    if not _ivr_campaign:
        return jsonify({"running": False, "processed": 0, "total": 0})
    return jsonify({
        "running":   _ivr_campaign.is_running,
        "processed": _ivr_campaign.processed,
        "total":     _ivr_campaign.total,
    })


@app.route("/ivr/results")
def ivr_results():
    """Descarga el CSV de resultados."""
    if not os.path.isfile(IVR_RESULTS_CSV):
        return jsonify({"ok": False, "error": "Sin resultados aún"}), 404
    from flask import send_file
    return send_file(IVR_RESULTS_CSV, as_attachment=True,
                     download_name="ivr_results.csv", mimetype="text/csv")


# ══════════════════════════════════════════════════════════════
#  MARCACIÓN MANUAL  —  flujo IVR completo
# ══════════════════════════════════════════════════════════════

_manual_call: "ManualCallSession | None" = None
_manual_lock = threading.Lock()


class ManualCallSession(threading.Thread):
    """
    Llamada manual con flujo IVR completo idéntico a IVRCampaign:
      - Marca vía ADB
      - Monitorea estado con CallMonitor
      - Registra _ivr_dtmf_callback (mismo mecanismo que la campaña)
      - Reproduce audio de bienvenida, menú IVR, despedida
      - Detecta tonos DTMF y registra la opción seleccionada
      - Emite manual_state y manual_log a la UI
    """

    def __init__(
        self,
        number:            str,
        device_id:         "str | None",
        audio_input:       "int | None",
        audio_output_idx:  "int | None",
        audio_output_name: "str | None",
        # Config IVR
        audio_welcome:     "str | None" = None,
        audio_menu:        "str | None" = None,
        audio_no_tone:     "str | None" = None,
        ivr_options:       "dict"       = None,
        tone_timeout:      int          = 10,
        menu_repeats:      int          = 2,
        record_calls:      bool         = False,
    ):
        super().__init__(daemon=True, name="ManualCall")
        self.number            = number
        self.device_id         = device_id
        self.audio_input       = audio_input
        self.audio_output_idx  = audio_output_idx
        self.audio_output_name = audio_output_name

        # IVR
        self.audio_welcome = audio_welcome
        self.audio_menu    = audio_menu
        self.audio_no_tone = audio_no_tone
        self.ivr_options   = ivr_options or {}
        self.tone_timeout  = tone_timeout
        self.menu_repeats  = menu_repeats
        self.record_calls  = record_calls

        self.state      = "IDLE"
        self.is_active  = False
        self._hangup_ev = threading.Event()
        # DTMF (mismo mecanismo que IVRCampaign)
        self._digit_event = threading.Event()
        self._last_digit: "str | None" = None

    # ── Helpers internos ──────────────────────────────────────────

    def _log(self, msg: str, level: str = "info"):
        print(f"[Manual] {msg}")
        socketio.emit("manual_log", {"msg": msg, "level": level, "ts": time.time()})

    def _set_state(self, state: str):
        self.state = state
        socketio.emit("manual_state", {"state": state, "number": self.number})

    def on_dtmf(self, digit: str):
        """Llamado por PythonAudioMonitor cuando detecta un tono DTMF."""
        print(f"[Manual] Dígito DTMF: {digit}")
        self._last_digit = digit
        self._digit_event.set()

    def _adb(self, *args: str):
        cmd = ["adb"]
        if self.device_id:
            cmd += ["-s", self.device_id]
        cmd += list(args)
        return subprocess.run(cmd, capture_output=True, timeout=10)

    def _hang_up(self):
        try:
            self._adb("shell", "input", "keyevent", "6")
        except Exception:
            pass

    def hangup(self):
        """Solicita colgar la llamada desde fuera del hilo."""
        self._hangup_ev.set()
        self._digit_event.set()   # desbloquear espera de tono

    def _gone(self, disconn_ev: threading.Event) -> bool:
        """True si el usuario colgó o la llamada se cortó."""
        return self._hangup_ev.is_set() or disconn_ev.is_set()

    # ── Bucle IVR (idéntico a IVRCampaign._handle_active) ────────

    def _run_ivr(self, disconn_ev: threading.Event):
        """
        Ejecuta el flujo IVR completo una vez la llamada está ACTIVE:
          1. Bienvenida
          2. Menú IVR (menu_repeats veces)
          3. Detección de tono válido
          4. Audio de despedida (global o por opción)
        """
        self._digit_event.clear()
        self._last_digit = None

        # --- 1. Bienvenida ---
        if self.audio_welcome:
            self._log("🎙️ Reproduciendo bienvenida...", "info")
            _play_audio(self.audio_welcome, cancel_event=self._hangup_ev)

        if self._gone(disconn_ev):
            return "DISCONNECTED_DURING_CALL", None

        # --- 2. Menú IVR ---
        for attempt in range(self.menu_repeats):
            if self._gone(disconn_ev):
                return "DISCONNECTED_DURING_CALL", None

            if self.audio_menu:
                self._log(f"📋 Menú IVR (intento {attempt+1}/{self.menu_repeats})...", "info")
                _play_audio(self.audio_menu, cancel_event=self._hangup_ev)

            if self._gone(disconn_ev):
                return "DISCONNECTED_DURING_CALL", None

            # Esperar tono válido
            self._log(f"⏳ Esperando tono DTMF ({self.tone_timeout}s)...", "info")
            deadline = time.time() + self.tone_timeout

            while time.time() < deadline:
                if self._gone(disconn_ev):
                    # Si ya hay dígito válido guardado, registrarlo
                    if self._last_digit and self._last_digit in self.ivr_options:
                        digit = self._last_digit
                        self._log(f"✅ Tono {digit!r} antes del cuelgue", "success")
                        return "ANSWERED_TONE", digit
                    return "DISCONNECTED_DURING_CALL", None

                remaining = max(0.05, deadline - time.time())
                self._digit_event.wait(timeout=min(remaining, 0.3))
                self._digit_event.clear()

                candidate = self._last_digit
                self._last_digit = None

                if candidate is None:
                    continue

                if candidate in self.ivr_options:
                    # Tono válido detectado
                    opt = self.ivr_options[candidate]
                    desc = opt.get("desc", candidate) if isinstance(opt, dict) else str(opt)
                    bye_path = opt.get("audio_bye") if isinstance(opt, dict) else None

                    self._log(f"🎯 Opción {candidate!r}: {desc}", "success")
                    socketio.emit("ivr_digit", {
                        "number": self.number, "digit": candidate, "option": desc
                    })

                    # Audio de despedida
                    bye = bye_path or self.audio_no_tone
                    if bye:
                        _play_audio(bye, cancel_event=self._hangup_ev)

                    return "ANSWERED_TONE", candidate
                else:
                    self._log(f"⚠️ Tono {candidate!r} no configurado — ignorado", "warn")

        # Sin respuesta tras todos los intentos
        if self.audio_no_tone:
            self._log("📢 Sin tono — reproduciendo audio de no respuesta...", "warn")
            _play_audio(self.audio_no_tone, cancel_event=self._hangup_ev)

        return "ANSWERED_NO_TONE", None

    # ── Hilo principal ────────────────────────────────────────────

    def run(self):
        global _audio_monitor_device, _audio_output_device_name, \
               _audio_output_device_index, _ivr_dtmf_callback

        # Convertir a int de forma segura
        audio_in  = int(self.audio_input)      if self.audio_input      is not None else None
        audio_out = int(self.audio_output_idx) if self.audio_output_idx is not None else None

        # Registrar globals de audio (igual que IVRCampaign)
        _audio_monitor_device      = audio_in
        _audio_output_device_index = audio_out
        _audio_output_device_name  = self.audio_output_name

        # Registrar callback DTMF (IGUAL que IVRCampaign.run())
        _ivr_dtmf_callback = self.on_dtmf

        # Inicializar mixer pygame
        if _PYGAME_OK:
            _ensure_mixer(self.audio_output_name)

        self.is_active = True
        self._set_state("DIALING")
        self._log(f"📞 Marcando: {self.number}", "info")

        # ─ 1. Marcar vía ADB ───────────────────────────────────────
        try:
            self._adb("shell", "am", "start", "-a",
                      "android.intent.action.CALL", "-d", f"tel:{self.number}")
        except Exception as exc:
            self._log(f"❌ Error ADB al marcar: {exc}", "error")
            self._set_state("ERROR")
            self.is_active = False
            _ivr_dtmf_callback = None
            return

        # ─ 2. Monitor de estado vía logcat ────────────────────────
        call_active_ev = threading.Event()
        disconn_ev     = threading.Event()
        monitor_stop   = threading.Event()

        def on_state(state: str):
            self._log(f"  → Estado: {state}", "info")
            if state in ("DIALING", "CONNECTING", "RINGING"):
                self._set_state("DIALING")
            elif state == "ACTIVE":
                call_active_ev.set()
                self._set_state("ACTIVE")
            elif state in ("DISCONNECTED", "TIMEOUT"):
                disconn_ev.set()
                call_active_ev.set()
                self._set_state("ENDED")

        if _MONITOR_OK:
            monitor = CallMonitor(device_id=self.device_id, timeout_s=CALL_MONITOR_TIMEOUT_S)
            monitor.start(on_state_change=on_state, stop_event=monitor_stop)
        else:
            self._log("⚠️ CallMonitor no disponible — esperando conexión...", "warn")
            monitor = None

        # Esperar ACTIVE o hangup (timeout 90s)
        deadline = time.time() + IVR_DIAL_TIMEOUT
        while not call_active_ev.is_set() and not self._hangup_ev.is_set():
            if time.time() > deadline:
                self._log("❌ Timeout esperando conexión", "error")
                self._set_state("ERROR")
                self._hang_up()
                break
            time.sleep(0.3)

        if self._hangup_ev.is_set() and not call_active_ev.is_set():
            self._log("🔴 Llamada cancelada por el usuario", "warn")
            self._hang_up()
            monitor_stop.set()
            self._set_state("ENDED")
            self.is_active = False
            _ivr_dtmf_callback = None
            return

        if disconn_ev.is_set():
            self._log("📥 Llamada terminada antes de contestar", "warn")
            monitor_stop.set()
            self._set_state("ENDED")
            self.is_active = False
            _ivr_dtmf_callback = None
            return

        # ─ 3. ACTIVE — iniciar monitor DTMF + flujo IVR ────────────────
        self._log("✅ Llamada contestada — iniciando flujo IVR...", "success")
        start_audio_monitor(audio_in)

        # Ejecutar flujo IVR completo
        result_status, result_digit = self._run_ivr(disconn_ev)

        # ─ 4. Colgar y limpiar ───────────────────────────────────
        monitor_stop.set()
        stop_audio_monitor()
        _ivr_dtmf_callback = None

        if self._hangup_ev.is_set():
            self._log("🔴 Colgando llamada...", "warn")
        self._hang_up()

        self._set_state("ENDED")
        self.is_active = False

        status_emoji = {
            "ANSWERED_TONE":     "✅",
            "ANSWERED_NO_TONE":  "⚠️",
            "DISCONNECTED_DURING_CALL": "📥",
        }.get(result_status, "ℹ️")
        self._log(
            f"{status_emoji} Resultado: {result_status}"
            + (f" — opción {result_digit}" if result_digit else ""),
            "success" if result_status == "ANSWERED_TONE" else "info"
        )
        self._log("✓ Llamada manual finalizada", "info")


# ── Rutas de marcación manual ───────────────────────────────────────────

@app.route("/ivr/manual/dial", methods=["POST"])
def manual_dial():
    """Inicia una llamada manual con flujo IVR completo."""
    global _manual_call

    # Guard 1: campaña activa
    with _ivr_lock:
        if _ivr_campaign and _ivr_campaign.is_running:
            return jsonify({"ok": False,
                            "error": "Hay una campaña activa. Deténla primero."}), 409

    # Guard 2: llamada manual ya activa
    with _manual_lock:
        if _manual_call and _manual_call.is_active:
            return jsonify({"ok": False,
                            "error": "Ya hay una llamada manual en curso."}), 409

    data   = request.get_json(force=True) or {}
    number = str(data.get("number", "")).strip()

    digits_only = number.replace("+", "").replace("-", "").replace(" ", "")
    if not number or not digits_only.isdigit() or len(digits_only) < 6:
        return jsonify({"ok": False, "error": "Número inválido (mínimo 6 dígitos)"}), 400

    device_id = data.get("device_id") or None
    if not device_id:
        return jsonify({"ok": False, "error": "Selecciona un dispositivo ADB"}), 400

    # ─ Dispositivos de audio ───────────────────────────────────────
    def _si(v):   # safe int
        try: return int(v) if v is not None and str(v).strip() != "" else None
        except: return None

    audio_in_idx  = _si(data.get("audio_device"))
    audio_out_idx = _si(data.get("audio_output_device"))
    audio_out_name: "str | None" = None

    if audio_out_idx is not None and _SD_OK:
        try:
            audio_out_name = sd.query_devices(audio_out_idx)["name"]
        except Exception:
            pass

    # ─ Config IVR ───────────────────────────────────────────────
    ivr_options_raw = data.get("ivr_options", {})
    # Normalizar: cada valor puede ser str o dict {desc, audio_bye}
    ivr_options: dict = {}
    if isinstance(ivr_options_raw, dict):
        for k, v in ivr_options_raw.items():
            if isinstance(v, dict):
                ivr_options[str(k)] = v
            else:
                ivr_options[str(k)] = str(v) if v else ""

    with _manual_lock:
        _manual_call = ManualCallSession(
            number            = number,
            device_id         = device_id,
            audio_input       = audio_in_idx,
            audio_output_idx  = audio_out_idx,
            audio_output_name = audio_out_name,
            audio_welcome     = data.get("audio_welcome") or None,
            audio_menu        = data.get("audio_menu")    or None,
            audio_no_tone     = data.get("audio_no_tone") or None,
            ivr_options       = ivr_options,
            tone_timeout      = int(data.get("tone_timeout", 10)),
            menu_repeats      = int(data.get("menu_repeats", 2)),
            record_calls      = bool(data.get("record_calls", False)),
        )
        _manual_call.start()

    return jsonify({"ok": True, "number": number})


@app.route("/ivr/manual/hangup", methods=["POST"])
def manual_hangup():
    """Cuelga la llamada manual activa."""
    with _manual_lock:
        if not _manual_call or not _manual_call.is_active:
            return jsonify({"ok": False, "error": "No hay llamada manual activa"}), 404
        _manual_call.hangup()
    return jsonify({"ok": True})


@app.route("/ivr/manual/status")
def manual_status():
    """Retorna el estado actual de la llamada manual."""
    with _manual_lock:
        if not _manual_call:
            return jsonify({"active": False, "state": "IDLE", "number": None})
        return jsonify({
            "active": _manual_call.is_active,
            "state":  _manual_call.state,
            "number": _manual_call.number,
        })


# ══════════════════════════════════════════════════════════════
#  PUENTE DE AUDIO  —  /ivr/bridge/*
# ══════════════════════════════════════════════════════════════

_active_bridge: "AudioBridge | None" = None
_bridge_lock    = threading.Lock()


def _bridge_on_status(msg: str, level: str):
    """Callback de estado del puente → emite a la UI vía Socket.IO."""
    print(f"[Bridge] {msg}")
    socketio.emit("bridge_log", {"msg": msg, "level": level, "ts": time.time()})


@app.route("/ivr/bridge/start", methods=["POST"])
def bridge_start():
    """
    Inicia el puente de audio bidireccional.
    Body JSON:
      phone_in_idx    int | null  — entrada que recibe audio del teléfono
      phone_out_idx   int | null  — salida que envía audio al teléfono
      pc_speaker_idx  int | null  — auriculares del agente (null = default)
      pc_mic_idx      int | null  — micrófono del agente   (null = default)
      block_ms        int         — tamaño de bloque en ms (default 40)
      gain_in         float       — ganancia teléfono→auriculares
      gain_out        float       — ganancia mic→teléfono
    """
    global _active_bridge
    with _bridge_lock:
        if _active_bridge and _active_bridge.is_running:
            return jsonify({"ok": False, "error": "Ya hay un puente activo"}), 409

    data = request.get_json(force=True) or {}

    phone_in   = data.get("phone_in_idx")
    phone_out  = data.get("phone_out_idx")
    pc_spk     = data.get("pc_speaker_idx")
    pc_mic     = data.get("pc_mic_idx")
    block_ms   = int(data.get("block_ms",  BRIDGE_BLOCK_MS))
    gain_in    = float(data.get("gain_in",  BRIDGE_GAIN_IN))
    gain_out   = float(data.get("gain_out", BRIDGE_GAIN_OUT))

    if phone_in is None and phone_out is None:
        return jsonify({
            "ok":    False,
            "error": "Debes seleccionar al menos la interfaz de audio del teléfono.",
        }), 400

    with _bridge_lock:
        _active_bridge = AudioBridge(
            phone_in_idx   = int(phone_in)  if phone_in  is not None else None,
            phone_out_idx  = int(phone_out) if phone_out is not None else None,
            pc_speaker_idx = int(pc_spk)    if pc_spk    is not None else None,
            pc_mic_idx     = int(pc_mic)    if pc_mic    is not None else None,
            block_ms       = block_ms,
            gain_in        = gain_in,
            gain_out       = gain_out,
            on_status      = _bridge_on_status,
        )
        _active_bridge.start()

    socketio.emit("bridge_state", {"state": "ACTIVE"})
    return jsonify({"ok": True, "msg": "Puente de audio iniciado"})


@app.route("/ivr/bridge/stop", methods=["POST"])
def bridge_stop():
    """Detiene el puente de audio activo."""
    global _active_bridge
    with _bridge_lock:
        if not _active_bridge or not _active_bridge.is_running:
            return jsonify({"ok": False, "error": "No hay puente activo"}), 404
        _active_bridge.stop()
        _active_bridge.join(timeout=3.0)
        _active_bridge = None

    socketio.emit("bridge_state", {"state": "IDLE"})
    return jsonify({"ok": True, "msg": "Puente detenido"})


@app.route("/ivr/bridge/status")
def bridge_status():
    """Estado del puente de audio."""
    with _bridge_lock:
        if not _active_bridge:
            return jsonify({"running": False, "error": None})
        return jsonify({
            "running": _active_bridge.is_running,
            "error":   _active_bridge.error,
        })


# ══════════════════════════════════════════════════════════════
#  PLANTILLAS  —  /ivr/templates/*
# ══════════════════════════════════════════════════════════════

@app.route("/ivr/templates", methods=["GET"])
def templates_list():
    """Lista todas las plantillas guardadas."""
    try:
        items = template_manager.list_all()
        return jsonify({"ok": True, "templates": items})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/ivr/templates", methods=["POST"])
def templates_save():
    """
    Guarda una plantilla nueva o actualiza una existente.
    Body JSON: { name: str, ...config }
    Los audios referenciados se copian al directorio templates/audio/.
    """
    data = request.get_json(force=True) or {}
    name = str(data.pop("name", "")).strip()
    if not name:
        return jsonify({"ok": False, "error": "El nombre de la plantilla es obligatorio"}), 400
    try:
        slug = template_manager.save(name, data)
        return jsonify({"ok": True, "slug": slug, "name": name})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/ivr/templates/<slug>", methods=["GET"])
def templates_load(slug):
    """Carga una plantilla por slug."""
    try:
        data = template_manager.load(slug)
        return jsonify({"ok": True, "template": data})
    except FileNotFoundError:
        return jsonify({"ok": False, "error": f"Plantilla '{slug}' no encontrada"}), 404
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/ivr/templates/<slug>", methods=["DELETE"])
def templates_delete(slug):
    """Elimina una plantilla y sus audios exclusivos."""
    try:
        deleted = template_manager.delete(slug)
        if deleted:
            return jsonify({"ok": True, "msg": f"Plantilla '{slug}' eliminada"})
        return jsonify({"ok": False, "error": "Plantilla no encontrada"}), 404
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/ivr/templates/<slug>", methods=["PUT"])
def templates_rename(slug):
    """Renombra una plantilla. Body JSON: { name: str }"""
    data     = request.get_json(force=True) or {}
    new_name = str(data.get("name", "")).strip()
    if not new_name:
        return jsonify({"ok": False, "error": "Nuevo nombre obligatorio"}), 400
    try:
        new_slug = template_manager.rename(slug, new_name)
        return jsonify({"ok": True, "slug": new_slug, "name": new_name})
    except FileNotFoundError:
        return jsonify({"ok": False, "error": f"Plantilla '{slug}' no encontrada"}), 404
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


# Servir audios de plantillas desde el directorio templates/audio/
@app.route("/ivr/templates/audio/<filename>")
def templates_audio_file(filename):
    """Sirve un archivo de audio guardado en la plantilla."""
    from flask import send_from_directory
    safe = os.path.basename(filename)
    if not os.path.isfile(os.path.join(TMPL_AUDIO, safe)):
        return jsonify({"ok": False, "error": "Archivo no encontrado"}), 404
    return send_from_directory(TMPL_AUDIO, safe)


# ══════════════════════════════════════════════════════════════════════
#  MULTI-SESSION — Rutas /ivr/sessions/*
#  Permite lanzar N automatizaciones simultáneas, cada una con su propio
#  dispositivo ADB y canal de audio.
# ══════════════════════════════════════════════════════════════════════

from dtmf_app.core.session_manager import session_manager
from dtmf_app.core.ivr_session     import SessionConfig
from dtmf_app.core.audio_probe     import AudioChannelProber
from dtmf_app.core.audio_player    import AudioPlayer

# ── Registro de sondas activas (probe en curso por device_id) ─────────────
_active_probes: "dict[str, AudioChannelProber]" = {}
_probes_lock   = threading.Lock()


def _session_emit(event: str, data: dict):
    """Wrapper para emitir eventos Socket.IO desde sesiones."""
    socketio.emit(event, data)


# ── Crear sesión ──────────────────────────────────────────────────────────

@app.route("/ivr/sessions", methods=["POST"])
def session_create():
    """
    Crea una nueva sesión de automatización.
    Body JSON (igual que /ivr/start más campos de sesión):
      device_id, audio_in_idx, audio_out_idx,
      numbers, delay_seconds, audio_welcome, audio_menu,
      audio_bye, audio_no_tone, ivr_options,
      tone_timeout, menu_repeats, record_calls, is_test,
      label (opcional)
    """
    data = request.get_json(force=True) or {}

    def _si(v):
        try: return int(v) if v is not None and str(v).strip() != "" else None
        except: return None

    audio_out_idx = _si(data.get("audio_out_idx") or data.get("audio_output_device"))
    audio_in_idx  = _si(data.get("audio_in_idx")  or data.get("audio_device"))

    # Nombre del dispositivo de salida
    audio_out_name = None
    if audio_out_idx is not None and _SD_OK:
        try:
            audio_out_name = sd.query_devices(audio_out_idx)["name"]
        except Exception:
            pass

    # Normalizar ivr_options
    ivr_options_raw = data.get("ivr_options", {})
    ivr_options: dict = {}
    if isinstance(ivr_options_raw, dict):
        for k, v in ivr_options_raw.items():
            ivr_options[str(k)] = v if isinstance(v, dict) else (str(v) if v else "")

    config = SessionConfig(
        device_id      = data.get("device_id", ""),
        audio_in_idx   = audio_in_idx,
        audio_out_idx  = audio_out_idx,
        audio_out_name = audio_out_name,
        numbers        = data.get("numbers", []),
        delay_seconds  = float(data.get("delay_seconds", 5)),
        audio_welcome  = data.get("audio_welcome") or None,
        audio_menu     = data.get("audio_menu")    or None,
        audio_bye      = data.get("audio_bye")     or None,
        audio_no_tone  = data.get("audio_no_tone") or None,
        ivr_options    = ivr_options,
        tone_timeout   = float(data.get("tone_timeout", 10)),
        menu_repeats   = int(data.get("menu_repeats", 2)),
        record_calls   = bool(data.get("record_calls", False)),
        is_test        = bool(data.get("is_test", False)),
    )

    # Validar conflictos con sesiones activas
    errors = session_manager.validate_config(config)
    if errors:
        return jsonify({"ok": False, "errors": errors}), 409

    session_id = session_manager.create(
        config   = config,
        emit_fn  = _session_emit,
        label    = data.get("label") or f"Sesión {data.get('device_id', '?')}",
    )

    return jsonify({"ok": True, "session_id": session_id,
                    "session": session_manager.get(session_id).to_dict()})


# ── Listar sesiones ───────────────────────────────────────────────────────

@app.route("/ivr/sessions", methods=["GET"])
def session_list():
    """Lista todas las sesiones registradas."""
    return jsonify({"ok": True, "sessions": session_manager.list_all()})


# ── Detalle de una sesión ─────────────────────────────────────────────────

@app.route("/ivr/sessions/<session_id>", methods=["GET"])
def session_get(session_id: str):
    session = session_manager.get(session_id)
    if session is None:
        return jsonify({"ok": False, "error": "Sesión no encontrada"}), 404
    return jsonify({"ok": True, "session": session.to_dict(),
                    "log": session.get_log()[-50:]})  # últimas 50 líneas


# ── Eliminar sesión ───────────────────────────────────────────────────────

@app.route("/ivr/sessions/<session_id>", methods=["DELETE"])
def session_delete(session_id: str):
    ok = session_manager.remove(session_id)
    if not ok:
        return jsonify({"ok": False,
                        "error": "Sesión no encontrada o está activa"}), 400
    return jsonify({"ok": True})


# ── Iniciar campaña de una sesión ─────────────────────────────────────────

@app.route("/ivr/sessions/<session_id>/start", methods=["POST"])
def session_start(session_id: str):
    """
    Inicia la campaña IVR de la sesión.
    Cada sesión usa su propio AudioPlayer y PythonAudioMonitor aislados.
    """
    session = session_manager.get(session_id)
    if session is None:
        return jsonify({"ok": False, "error": "Sesión no encontrada"}), 404

    cfg = session.config

    if not cfg.device_id:
        return jsonify({"ok": False, "error": "Sesión sin device_id"}), 400
    if not cfg.numbers:
        return jsonify({"ok": False, "error": "Sin números en la cola"}), 400

    # Validar conflictos
    errors = session_manager.validate_config(cfg)
    if errors:
        return jsonify({"ok": False, "errors": errors}), 409

    # Construir AudioPlayer dedicado para esta sesión
    player = AudioPlayer(output_device_idx=cfg.audio_out_idx)

    # Construir PythonAudioMonitor dedicado (con callback aislado)
    # El callback lo registrará la campaña cuando se inicie
    def _campaign_factory(sess):
        """Crea y configura un IVRCampaign para esta sesión."""
        campaign_cfg = cfg.to_campaign_config()
        campaign     = IVRCampaign(campaign_cfg)

        # Sobreescribir métodos de audio para usar el AudioPlayer de sesión
        # y el PythonAudioMonitor con session_id
        original_run = campaign.run

        def patched_run():
            """Run con dispositivos de audio aislados por sesión."""
            # Configurar monitor DTMF con callback de la campaña
            monitor = PythonAudioMonitor(
                device_index    = cfg.audio_in_idx,
                dtmf_callback   = campaign.on_dtmf,
                session_id      = sess.session_id,
            )

            # Patch de funciones de audio en el contexto de esta campaña
            # Usamos atributos de campaña para evitar tocar globals
            campaign._session_player  = player
            campaign._session_monitor = monitor
            campaign._session_id      = sess.session_id

            # Emitir progreso a la UI con session_id
            def _session_emit_ivr(event, data):
                data["session_id"] = sess.session_id
                socketio.emit(event, data)

            # El IVRCampaign usa _emit_ivr global — redirigir vía monkey-patch temporal
            # es complejo. En cambio, dejamos que los eventos globales lleguen a la UI
            # y el frontend los filtra por session_id donde aplique.
            # Los logs de sesión se emiten también via session_log.
            monitor.start()
            original_run()
            monitor.stop()

            sess.set_status("DONE")

        import types
        campaign.run = types.MethodType(lambda self: patched_run(), campaign)
        return campaign

    ok = session_manager.start(session_id, _campaign_factory)
    if not ok:
        return jsonify({"ok": False, "error": "Error iniciando campaña"}), 500

    # Iniciar watchdog ADB para esta sesión
    watchdog = ADBWatchdog(cfg.device_id)
    watchdog.start()
    session.watchdog = watchdog

    return jsonify({"ok": True, "session_id": session_id})


# ── Detener sesión ────────────────────────────────────────────────────────

@app.route("/ivr/sessions/<session_id>/stop", methods=["POST"])
def session_stop(session_id: str):
    ok = session_manager.stop(session_id)
    return jsonify({"ok": ok, "error": "Sesión no encontrada" if not ok else None})


# ── Pausar / Reanudar ─────────────────────────────────────────────────────

@app.route("/ivr/sessions/<session_id>/pause", methods=["POST"])
def session_pause(session_id: str):
    ok = session_manager.pause(session_id)
    return jsonify({"ok": ok})


@app.route("/ivr/sessions/<session_id>/resume", methods=["POST"])
def session_resume(session_id: str):
    ok = session_manager.resume(session_id)
    return jsonify({"ok": ok})


# ── Actualizar config de sesión (sin reiniciar) ────────────────────────────

@app.route("/ivr/sessions/<session_id>", methods=["PATCH"])
def session_update(session_id: str):
    """Actualiza campos de config de una sesión que aún no ha iniciado."""
    session = session_manager.get(session_id)
    if session is None:
        return jsonify({"ok": False, "error": "Sesión no encontrada"}), 404
    if session.is_active:
        return jsonify({"ok": False, "error": "No se puede editar una sesión activa"}), 409

    data = request.get_json(force=True) or {}
    cfg  = session.config

    def _si(v):
        try: return int(v) if v is not None and str(v).strip() != "" else None
        except: return None

    if "numbers"       in data: cfg.numbers       = data["numbers"]
    if "delay_seconds" in data: cfg.delay_seconds = float(data["delay_seconds"])
    if "audio_welcome" in data: cfg.audio_welcome = data["audio_welcome"] or None
    if "audio_menu"    in data: cfg.audio_menu    = data["audio_menu"]    or None
    if "audio_bye"     in data: cfg.audio_bye     = data["audio_bye"]     or None
    if "audio_no_tone" in data: cfg.audio_no_tone = data["audio_no_tone"] or None
    if "ivr_options"   in data: cfg.ivr_options   = data["ivr_options"]
    if "tone_timeout"  in data: cfg.tone_timeout  = float(data["tone_timeout"])
    if "menu_repeats"  in data: cfg.menu_repeats  = int(data["menu_repeats"])
    if "record_calls"  in data: cfg.record_calls  = bool(data["record_calls"])
    if "label"         in data: session.label      = data["label"]

    # Audio devices
    ao = _si(data.get("audio_out_idx") or data.get("audio_output_device"))
    ai = _si(data.get("audio_in_idx")  or data.get("audio_device"))
    if ao is not None:
        cfg.audio_out_idx = ao
        if _SD_OK:
            try: cfg.audio_out_name = sd.query_devices(ao)["name"]
            except: pass
    if ai is not None:
        cfg.audio_in_idx = ai

    return jsonify({"ok": True, "session": session.to_dict()})


# ══════════════════════════════════════════════════════════════════════
#  AUTO-DETECCIÓN DE CANAL DE AUDIO — /ivr/probe
# ══════════════════════════════════════════════════════════════════════

@app.route("/ivr/probe/start", methods=["POST"])
def probe_start():
    """
    Inicia la auto-detección de canal de audio para un dispositivo ADB.
    Body JSON: { device_id, session_id (opcional) }

    Proceso:
      1. Push de calib_tone.wav al dispositivo vía ADB
      2. Reproducción en el altavoz del teléfono
      3. Escucha en todos los canales libres buscando 3750 Hz (Goertzel)
      4. Emite 'probe_result' vía Socket.IO con el canal detectado
    """
    data      = request.get_json(force=True) or {}
    device_id = data.get("device_id", "").strip()
    session_id_hint = data.get("session_id")

    if not device_id:
        return jsonify({"ok": False, "error": "device_id requerido"}), 400

    with _probes_lock:
        if device_id in _active_probes:
            return jsonify({"ok": False,
                            "error": "Ya hay una sonda activa para este dispositivo"}), 409

    # Canales de entrada ya ocupados por sesiones activas
    occupied = session_manager.occupied_inputs()

    def _on_found(in_idx: int, out_idx: "int | None"):
        socketio.emit("probe_result", {
            "ok":        True,
            "device_id": device_id,
            "in_idx":    in_idx,
            "out_idx":   out_idx,
            "session_id": session_id_hint,
            "msg":       f"✅ Canal detectado: entrada={in_idx}" +
                         (f" salida={out_idx}" if out_idx is not None else ""),
        })
        with _probes_lock:
            _active_probes.pop(device_id, None)

        # Si hay session_id, actualizar automáticamente la sesión
        if session_id_hint:
            session = session_manager.get(session_id_hint)
            if session:
                session.config.audio_in_idx  = in_idx
                session.config.audio_out_idx = out_idx
                if out_idx is not None and _SD_OK:
                    try:
                        session.config.audio_out_name = sd.query_devices(out_idx)["name"]
                    except Exception:
                        pass
                session.log(f"🔊 Canales auto-detectados: in={in_idx} out={out_idx}", "success")

    def _on_error(msg: str):
        socketio.emit("probe_result", {
            "ok":        False,
            "device_id": device_id,
            "session_id": session_id_hint,
            "msg":       f"❌ {msg}",
        })
        with _probes_lock:
            _active_probes.pop(device_id, None)

    def _on_status(msg: str):
        socketio.emit("probe_status", {
            "device_id": device_id,
            "session_id": session_id_hint,
            "msg":       msg,
        })

    prober = AudioChannelProber(
        device_id        = device_id,
        occupied_inputs  = occupied,
        on_found         = _on_found,
        on_error         = _on_error,
        on_status        = _on_status,
    )

    with _probes_lock:
        _active_probes[device_id] = prober

    prober.start()

    return jsonify({
        "ok":      True,
        "msg":     "Sonda iniciada — espera el evento 'probe_result' vía Socket.IO",
        "device_id": device_id,
    })


@app.route("/ivr/probe/status/<device_id>")
def probe_status(device_id: str):
    """Devuelve si hay una sonda activa para el device_id dado."""
    with _probes_lock:
        active = device_id in _active_probes
    return jsonify({"active": active, "device_id": device_id})


if __name__ == "__main__":
    print("=" * 55)
    print("  DTMF Analyzer - Servidor Web  (WebSocket activo)")
    print("  Abre: http://localhost:5050")
    print("=" * 55)
    socketio.run(app, host="0.0.0.0", port=5050, debug=False, allow_unsafe_werkzeug=True)
